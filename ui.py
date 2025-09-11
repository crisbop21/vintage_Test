# app.py — QOB (quarters-on-book) version with:
# - Chart: X axis in MONTHS (QOB*3), Y axis in %
# - Vintage table: nicer aesthetics, PD as %
# - Fast, vectorized pipeline + optional Numba speed-up
# - Progress bar while building curves
# - Integrity checks + summary-only PDF + Excel samples

import os
os.environ["NUMEXPR_MAX_THREADS"] = "8"

import math
import hashlib
import textwrap
import re
from io import BytesIO
from typing import Optional, Callable

import numpy as np
import pandas as pd
import matplotlib
matplotlib.rcParams["path.simplify"] = True
matplotlib.rcParams["agg.path.chunksize"] = 10000
matplotlib.rcParams["axes.unicode_minus"] = False
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import plotly.graph_objects as go
from plotly.colors import hex_to_rgb, qualitative, sequential
import plotly.io as pio

import streamlit as st
from data_processing import load_full, build_chart_data_fast_quarter, compute_vintage_default_summary
from integrity_checks import run_integrity_checks, export_integrity_pdf, export_issues_excel
from plotting import plot_curves_percent_with_months

# Progress helper
def mk_progress_updater(bar, steps: int = 5) -> Callable[[str], None]:
    ctr = {'i': 0, 'steps': max(1, int(steps))}
    def _update(msg: str):
        ctr['i'] += 1
        bar.progress(min(ctr['i']/ctr['steps'], 1.0), text=msg)
    return _update

st.set_page_config(page_title='Vintage Curves (QOB) + Integrity — Ultra-Fast', layout='wide')

# Density selector allows comfortable or compact spacing modes
density_mode = st.sidebar.selectbox('Density', ['Comfortable', 'Compact'], index=0)

# Blue palette with white background
# Darker blues for sidebar and accents
WB_PRIMARY = "#1E3A8A"      # Button and link color
WB_SECONDARY = "#1E40AF"    # Sidebar background
WB_BG = "#FFFFFF"           # Page background
WB_TEXT = "#002244"         # General text color
st.markdown(
    f"""
    <style>
        :root {{
            --space-1: 8px;
            --space-2: 16px;
            --space-3: 24px;
            --space-4: 32px;
            --space-5: 48px;
        }}
        html, body, [data-testid="stAppViewContainer"], .stApp {{
            background-color: {WB_BG};
            line-height: 1.5;
        }}
        [data-testid="stAppViewContainer"] {{
            padding: var(--space-3);
        }}
        *, *::placeholder {{
            color: {WB_TEXT} !important;
        }}
       [data-testid="stSidebar"] {{
            background-color: {WB_SECONDARY};
            padding: var(--space-3);
        }}
        [data-testid="stSidebar"] * {{
            color: white !important;
        }}
        [data-testid="stSidebar"] .stSelectbox * {{
            color: black !important;
        }}
        .block-container {{
            padding-top: var(--space-4);
            padding-bottom: var(--space-4);
        }}
        .stButton>button {{
            background-color: {WB_PRIMARY};
            margin-top: var(--space-2);
            margin-bottom: var(--space-2);
            padding: var(--space-1) var(--space-2);
            border: none;
        }}
        .stButton>button, .stButton>button * {{
            color: white !important;
        }}
        .stButton>button:hover {{
            filter: brightness(1.1);
        }}
        .stButton>button:active {{
            filter: brightness(0.9);
        }}
        .stButton>button:focus {{
            outline: 3px solid white;
            outline-offset: 2px;
        }}
        a {{
            color: {WB_PRIMARY} !important;
        }}
         p {{
            max-width: 75ch;
        }}
        [data-baseweb="tag"] {{
            background-color: {WB_PRIMARY};
            color: white !important;
        }}
         p {{
            max-width: 75ch;
        }}
        input[type="checkbox"], input[type="radio"], input[type="range"] {{
            accent-color: {WB_PRIMARY};
        }}
        [data-baseweb="tag"] {{
            background-color: {WB_PRIMARY};
        }}
        [data-baseweb="tag"], [data-baseweb="tag"] * {{
            color: white !important;
        }}
        [data-baseweb="tag"] [data-baseweb="close"] svg {{
            fill: white !important;
        }}
    </style>
    """,
    unsafe_allow_html=True,
)

# Compact density adjustments
if density_mode == 'Compact':
    st.markdown(
        """
        <style>
            [data-testid="stAppViewContainer"] {
                padding: var(--space-2);
            }
            .stButton>button {
                padding: var(--space-1);
                margin-top: var(--space-1);
                margin-bottom: var(--space-1);
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

st.markdown("<h1 style='text-align: center;'>Vintage Default-Rate Tool</h1>", unsafe_allow_html=True)

MAX_MB = 50
RESERVED_COLS = {
    'Loan ID','Origination date','Maturity date','Observation date',
    'Days past due','Origination amount','Current amount',
    'Vintage','MOB','QOB','is_def','is_def_cum'
}

# ──────────────────────────────────────────────────────────────────────────────
# Helpers: schema, typing, ageing
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header('Instructions')
    st.markdown(
        "1. Configure settings below.\n"
        "2. Upload an Excel (.xlsx) file.\n"
        "3. Click **Load dataset**.\n"
        "4. Explore the tabs for integrity checks, tables, and charts."
    )

left, right = st.columns([0.5, 2], gap="large")

with left:
    st.header('Settings')
    dpd_threshold = st.number_input('Default if Days past due ≥', min_value=1, max_value=365, value=90, step=1)
    pretty_ints = st.checkbox(
        'Thousands separators for integer columns',
        value=False,
        help='Shows 12,345 instead of 12345; disables numeric sorting on those two columns.',
    )

    st.subheader('Upload Excel')
    uploaded = st.file_uploader('Upload a .xlsx file', type=['xlsx'], accept_multiple_files=False)

    # Persist full dataset
    if 'df_full' not in st.session_state:
        st.session_state['df_full'] = None

    if uploaded:
        size_mb = uploaded.size / (1024 * 1024)
        st.caption(f'File size, {size_mb:,.1f} MB')
        if size_mb > MAX_MB:
            st.warning('Large file, consider filtering columns or using CSV.')

        from openpyxl import load_workbook
        names = load_workbook(filename=BytesIO(uploaded.getvalue()), read_only=True, data_only=True).sheetnames
        sheet = st.selectbox('Select sheet', options=names, index=0)
        header_row = st.number_input('Header row [1 means first row]', min_value=1, value=1, step=1)

        if st.button('Load dataset', type='primary'):
            with st.status('Loading dataset...', expanded=False) as status:
                df_full = load_full(uploaded.getvalue(), sheet=sheet, header=header_row - 1)
                st.session_state['df_full'] = df_full
                status.update(label='Dataset loaded.', state='complete')
with right:
    if st.session_state['df_full'] is not None:
        st.divider()
        chosen_df_raw = st.session_state['df_full']
        tab_integrity, tab_tables, tab_charts = st.tabs(["Integrity", "Tables", "Charts"])

        # Integrity
        with tab_integrity:
            st.subheader('Integrity checks, PDF (summary only) & Excel export')
            dataset_label = 'Full'

            if st.button('Run integrity checks and generate outputs'):
                with st.status('Running checks...', expanded=False):
                    summary, issues_df, vintage_issues_df = run_integrity_checks(chosen_df_raw, dpd_threshold=dpd_threshold)

                if 'fatal' in summary:
                    st.error(summary['fatal'])
                else:
                    st.success('Checks complete.')
                    st.json(summary)

                    pdf_bytes = export_integrity_pdf(summary, dataset_label=dataset_label)
                    st.download_button('Download integrity report (PDF, summary-only)', pdf_bytes,
                                       'integrity_report.pdf', 'application/pdf')

                    xlsx_bytes = export_issues_excel(issues_df, vintage_issues_df)
                    st.download_button('Download issues sample (Excel)', xlsx_bytes,
                                       'integrity_issues_sample.xlsx',
                                       'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                    if issues_df is not None and not issues_df.empty:
                        st.caption('Row-level issues (sample)')
                        st.dataframe(issues_df.head(200), use_container_width=True)
                    else:
                        st.info('No row-level issues sampled.')

                    if vintage_issues_df is not None and not vintage_issues_df.empty:
                        st.caption('Vintage/cohort issues')
                        st.dataframe(vintage_issues_df.head(200), use_container_width=True)
                    else:
                        st.info('No vintage-level issues detected.')

            st.divider()

        # ---- Vintage table (explicit formatting) ----
        with tab_tables:
            st.subheader('Unique loans & default summary by vintage')
            try:
                summary_df = compute_vintage_default_summary(chosen_df_raw, dpd_threshold=dpd_threshold)
                st.caption('Observation_Time = default date − first obs (if defaulted), else last obs − first obs (years).')

                # Rename only for display
                disp = summary_df.rename(columns={
                    "Unique_loans": "Unique loans",
                    "Defaulted_loans": "Defaulted loans",
                    "Observation_Time": "Obs Time (years)",
                    "Default_rate_pa": "Annualized default rate",
                    "Cum_PD": "Cum PD",
                })
                disp["Cum PD (%)"] = disp["Cum PD"] * 100
                disp["Annualized default rate (%)"] = disp["Annualized default rate"] * 100

                table = disp[[
                    "Vintage",
                    "Unique loans",
                    "Defaulted loans",
                    "Cum PD (%)",
                    "Obs Time (years)",
                    "Annualized default rate (%)",
                ]]
                styles = {
                    "Unique loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                    "Defaulted loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                    "Cum PD (%)": "{:.2f}",
                    "Obs Time (years)": "{:.2f}",
                    "Annualized default rate (%)": "{:.2f}",
                }
                styler = (
                    table.style
                    .format(styles)
                    .background_gradient(subset=["Cum PD (%)", "Annualized default rate (%)"], cmap="Reds")
                    .hide(axis="index")
                )

                st.dataframe(styler, use_container_width=True)

                st.download_button(
                    'Download CSV (vintage default summary)',
                    summary_df.to_csv(index=False).encode('utf-8'),
                    'vintage_default_summary.csv','text/csv'
                )
            except Exception as e:
                st.info(f'Could not compute vintage default summary: {e}')

            st.divider()

        # ---- Vintage curves — QOB engine, months axis, % y, legend ----
        with tab_charts:
            st.subheader('Vintage curves (months on axis)')
            col_chart, col_settings = st.columns([3, 0.5], gap="large")

            with col_settings:
                st.header('Chart settings')
                max_months_show = st.slider('Show curves up to (months)', min_value=12, max_value=180, value=60, step=6)
                show_legend = st.checkbox('Show legend in chart', value=True)
                palette_option = st.selectbox('Color palette', ['Gradient', 'Plotly', 'Viridis'])
                base_color = st.color_picker(
                    'Base chart color',
                    value=st.get_option("theme.primaryColor") or '#1f77b4',
                    help='Used when Gradient palette is selected.',
                )
                line_width = st.slider('Line width', min_value=1, max_value=5, value=1)

            with col_chart:
                try:
                    prog_bar = st.progress(0.0, text="Initializing …")
                    upd = mk_progress_updater(prog_bar, steps=5)

                    max_qob_show = max(1, math.ceil(max_months_show / 3))
                    df_plot_any = build_chart_data_fast_quarter(
                        chosen_df_raw, dpd_threshold=dpd_threshold, max_qob=max_qob_show, prog=upd
                    )

                    if df_plot_any.empty:
                        prog_bar.progress(1.0, text="No data to plot.")
                        st.info('Not enough data to plot curves for the chosen dataset.')
                    else:
                        vintages = df_plot_any.columns.tolist()
                        selected_vintages = st.multiselect('Vintages to display', vintages, default=vintages)
                        if not selected_vintages:
                            prog_bar.progress(1.0, text="No vintages selected.")
                            st.info('Select at least one vintage to plot.')
                        else:
                            df_plot = df_plot_any[selected_vintages]
                            prog_bar.progress(0.9, text="Rendering chart …")
                            ttl = f'Vintage Default-Rate Evolution | DPD≥{dpd_threshold}'
                            fig = plot_curves_percent_with_months(
                                df_wide=df_plot,
                                title=ttl,
                                show_legend=show_legend,
                                legend_limit=50,
                                palette=palette_option,
                                base_color=base_color,
                                line_width=line_width,
                            )
                            prog_bar.progress(1.0, text="Done")

                            st.plotly_chart(fig, use_container_width=True)

                            export_df = df_plot.reset_index().rename(columns={"QOB":"QOB"})
                            export_df.insert(0, "Months", export_df["QOB"] * 3)
                            st.download_button('Download curves (CSV; Months + QOB)',
                                               export_df.to_csv(index=False).encode('utf-8'),
                                               'vintage_curves_qob.csv','text/csv')

                except Exception as e:
                    st.info(f'Plot skipped, {e}')


    else:
        st.caption('Upload an Excel to continue.')
        














