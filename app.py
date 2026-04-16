# app.py — QOB (quarters-on-book) version with:
# - Chart: X axis in MONTHS (QOB*3), Y axis in %
# - Vintage table: nicer aesthetics, PD as %
# - Fast, vectorized pipeline + optional Numba speed-up
# - Progress bar while building curves
# - Integrity checks + summary-only PDF + Excel samples

import os
os.environ["NUMEXPR_MAX_THREADS"] = "8"

import gc
import logging
import math
import hashlib
import textwrap
import re
import traceback
from io import BytesIO
from typing import Optional, Callable

import numpy as np
import pandas as pd
import matplotlib
matplotlib.rcParams["path.simplify"] = True
matplotlib.rcParams["agg.path.chunksize"] = 10000
matplotlib.rcParams["axes.unicode_minus"] = False
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import plotly.graph_objects as go
from plotly.colors import hex_to_rgb, qualitative, sequential
import streamlit as st

# ──────────────────────────────────────────────────────────────────────────────
# Logging — prints to stdout which Streamlit Cloud captures in "Manage app" logs
# ──────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
    force=True,
)
logger = logging.getLogger("vintage_app")

def _log_resource_usage(label: str = ""):
    """Log memory usage so crashes leave a trail in Streamlit Cloud logs."""
    try:
        import resource
        mem_mb = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss / 1024  # Linux: KB → MB
        logger.info("MEMORY %s — RSS %.0f MB", label, mem_mb)
    except Exception:
        pass

_log_resource_usage("startup")

# ──────────────────────────────────────────────────────────────────────────────
# Smart cache decorator: avoids "missing ScriptRunContext … bare mode" spam
# ──────────────────────────────────────────────────────────────────────────────
def cache_data_smart(**kwargs):
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        if get_script_run_ctx() is None:
            def passthrough(func): return func
            return passthrough
    except Exception:
        pass
    return st.cache_data(**kwargs)

# Optional Numba acceleration for per-loan cummax
try:
    from numba import njit
    NUMBA_OK = True
except Exception:
    NUMBA_OK = False

st.set_page_config(
    page_title='Vintage Default-Rate Analytics | Corporate Dashboard',
    page_icon='📊',
    layout='wide',
    initial_sidebar_state='expanded'
)

# ──────────────────────────────────────────────────────────────────────────────
# Corporate Color Palette (keeping user's colors)
# ──────────────────────────────────────────────────────────────────────────────
WB_PRIMARY = "#1E3A8A"       # Primary blue - buttons, accents
WB_SECONDARY = "#1E40AF"     # Secondary blue - sidebar
WB_BG = "#FFFFFF"            # Background white
WB_TEXT = "#002244"          # Text color
WB_LIGHT = "#F8FAFC"         # Light background for cards
WB_BORDER = "#E2E8F0"        # Subtle border color
WB_ACCENT = "#3B82F6"        # Lighter accent blue
WB_SUCCESS = "#10B981"       # Success green
WB_MUTED = "#64748B"         # Muted text

# Density selector in sidebar
density_mode = st.sidebar.selectbox('Display Density', ['Comfortable', 'Compact'], index=0)

# ──────────────────────────────────────────────────────────────────────────────
# Corporate CSS Styling
# ──────────────────────────────────────────────────────────────────────────────
st.markdown(
    f"""
    <style>
        /* ═══════════════════════════════════════════════════════════════════
           CSS VARIABLES & BASE STYLES
           ═══════════════════════════════════════════════════════════════════ */
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

        :root {{
            --primary: {WB_PRIMARY};
            --secondary: {WB_SECONDARY};
            --bg: {WB_BG};
            --text: {WB_TEXT};
            --light: {WB_LIGHT};
            --border: {WB_BORDER};
            --accent: {WB_ACCENT};
            --success: {WB_SUCCESS};
            --muted: {WB_MUTED};
            --space-1: 8px;
            --space-2: 16px;
            --space-3: 24px;
            --space-4: 32px;
            --space-5: 48px;
            --radius-sm: 6px;
            --radius-md: 10px;
            --radius-lg: 16px;
            --shadow-sm: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
            --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
            --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
            --shadow-xl: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04);
            --transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
        }}

        /* ═══════════════════════════════════════════════════════════════════
           GLOBAL STYLES
           ═══════════════════════════════════════════════════════════════════ */
        html, body, [data-testid="stAppViewContainer"], .stApp {{
            background: linear-gradient(135deg, {WB_BG} 0%, {WB_LIGHT} 100%);
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6;
            -webkit-font-smoothing: antialiased;
        }}

        [data-testid="stAppViewContainer"] {{
            padding: var(--space-4);
        }}

        *, *::placeholder {{
            color: var(--text) !important;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           SIDEBAR STYLES - Light background with dark text for contrast
           ═══════════════════════════════════════════════════════════════════ */
        [data-testid="stSidebar"] {{
            background: {WB_LIGHT};
            padding: 0;
            box-shadow: var(--shadow-lg);
            border-right: 1px solid {WB_BORDER};
        }}

        [data-testid="stSidebar"] > div:first-child {{
            padding: var(--space-3);
        }}

        [data-testid="stSidebar"] * {{
            color: {WB_TEXT} !important;
        }}

        [data-testid="stSidebar"] .stSelectbox *,
        [data-testid="stSidebar"] [data-baseweb="select"] * {{
            color: {WB_TEXT} !important;
        }}

        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3 {{
            color: {WB_PRIMARY} !important;
            font-weight: 600;
            letter-spacing: -0.02em;
        }}

        [data-testid="stSidebar"] .stMarkdown p {{
            color: {WB_TEXT} !important;
            font-size: 0.9rem;
        }}

        /* Sidebar divider */
        [data-testid="stSidebar"] hr {{
            border: none;
            height: 1px;
            background: {WB_BORDER};
            margin: var(--space-3) 0;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           MAIN CONTENT AREA
           ═══════════════════════════════════════════════════════════════════ */
        .block-container {{
            padding: var(--space-4) var(--space-5);
            max-width: 1400px;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           HEADERS & TYPOGRAPHY
           ═══════════════════════════════════════════════════════════════════ */
        h1 {{
            font-weight: 700;
            letter-spacing: -0.03em;
            color: var(--text) !important;
        }}

        h2 {{
            font-weight: 600;
            letter-spacing: -0.02em;
            color: var(--text) !important;
            border-bottom: 2px solid var(--border);
            padding-bottom: var(--space-2);
            margin-bottom: var(--space-3);
        }}

        h3 {{
            font-weight: 600;
            color: var(--primary) !important;
            letter-spacing: -0.01em;
        }}

        p {{
            max-width: 75ch;
            color: var(--text);
        }}

        /* ═══════════════════════════════════════════════════════════════════
           BUTTONS
           ═══════════════════════════════════════════════════════════════════ */
        .stButton > button {{
            background: linear-gradient(135deg, {WB_PRIMARY} 0%, {WB_SECONDARY} 100%);
            color: white !important;
            border: none;
            border-radius: var(--radius-md);
            padding: 12px 24px;
            font-weight: 600;
            font-size: 0.9rem;
            letter-spacing: 0.02em;
            text-transform: uppercase;
            box-shadow: var(--shadow-md);
            transition: var(--transition);
            margin: var(--space-2) 0;
        }}

        .stButton > button:hover {{
            transform: translateY(-2px);
            box-shadow: var(--shadow-lg);
            filter: brightness(1.05);
        }}

        .stButton > button:active {{
            transform: translateY(0);
            box-shadow: var(--shadow-sm);
        }}

        .stButton > button:focus {{
            outline: 3px solid {WB_ACCENT};
            outline-offset: 2px;
        }}

        .stButton > button, .stButton > button * {{
            color: white !important;
        }}

        /* Secondary/Download buttons */
        .stDownloadButton > button {{
            background: var(--light);
            color: var(--primary) !important;
            border: 2px solid var(--border);
            border-radius: var(--radius-md);
            padding: 10px 20px;
            font-weight: 500;
            transition: var(--transition);
        }}

        .stDownloadButton > button:hover {{
            background: var(--primary);
            color: white !important;
            border-color: var(--primary);
        }}

        .stDownloadButton > button * {{
            transition: var(--transition);
        }}

        .stDownloadButton > button:hover * {{
            color: white !important;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           FORM ELEMENTS
           ═══════════════════════════════════════════════════════════════════ */
        /* Text inputs */
        .stTextInput > div > div > input,
        .stNumberInput > div > div > input {{
            border: 2px solid var(--border);
            border-radius: var(--radius-sm);
            padding: 10px 14px;
            font-size: 0.95rem;
            transition: var(--transition);
            background: white;
        }}

        .stTextInput > div > div > input:focus,
        .stNumberInput > div > div > input:focus {{
            border-color: var(--primary);
            box-shadow: 0 0 0 3px rgba(30, 58, 138, 0.1);
        }}

        /* Select boxes */
        [data-baseweb="select"] {{
            border-radius: var(--radius-sm) !important;
        }}

        [data-baseweb="select"] > div {{
            border: 2px solid var(--border) !important;
            border-radius: var(--radius-sm) !important;
            background: white !important;
            transition: var(--transition);
        }}

        [data-baseweb="select"] > div:hover {{
            border-color: var(--primary) !important;
        }}

        /* Checkboxes and radios */
        input[type="checkbox"], input[type="radio"] {{
            accent-color: var(--primary);
            width: 18px;
            height: 18px;
        }}

        input[type="range"] {{
            accent-color: var(--primary);
        }}

        /* File uploader */
        [data-testid="stFileUploader"] {{
            border: 2px dashed var(--border);
            border-radius: var(--radius-lg);
            padding: var(--space-3);
            background: var(--light);
            transition: var(--transition);
        }}

        [data-testid="stFileUploader"]:hover {{
            border-color: var(--primary);
            background: white;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           TABS
           ═══════════════════════════════════════════════════════════════════ */
        .stTabs [data-baseweb="tab-list"] {{
            background: var(--light);
            border-radius: var(--radius-lg);
            padding: 6px;
            gap: 8px;
        }}

        .stTabs [data-baseweb="tab"] {{
            background: transparent;
            border-radius: var(--radius-md);
            padding: 12px 24px;
            font-weight: 500;
            color: var(--muted) !important;
            transition: var(--transition);
        }}

        .stTabs [data-baseweb="tab"]:hover {{
            background: white;
            color: var(--text) !important;
        }}

        .stTabs [aria-selected="true"] {{
            background: white !important;
            color: var(--primary) !important;
            box-shadow: var(--shadow-sm);
        }}

        .stTabs [data-baseweb="tab-highlight"] {{
            display: none;
        }}

        .stTabs [data-baseweb="tab-border"] {{
            display: none;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           DATA TABLES
           ═══════════════════════════════════════════════════════════════════ */
        [data-testid="stDataFrame"] {{
            border: 1px solid var(--border);
            border-radius: var(--radius-lg);
            overflow: hidden;
            box-shadow: var(--shadow-sm);
        }}

        [data-testid="stDataFrame"] table {{
            font-size: 0.9rem;
        }}

        [data-testid="stDataFrame"] th {{
            background: var(--light) !important;
            color: var(--text) !important;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.8rem;
            letter-spacing: 0.05em;
            padding: 14px 16px !important;
        }}

        [data-testid="stDataFrame"] td {{
            padding: 12px 16px !important;
            border-bottom: 1px solid var(--border);
        }}

        [data-testid="stDataFrame"] tr:hover td {{
            background: var(--light) !important;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           CARDS & CONTAINERS
           ═══════════════════════════════════════════════════════════════════ */
        [data-testid="stExpander"] {{
            background: white;
            border: 1px solid var(--border);
            border-radius: var(--radius-lg);
            box-shadow: var(--shadow-sm);
            overflow: hidden;
        }}

        [data-testid="stExpander"] summary {{
            font-weight: 600;
            padding: var(--space-2) var(--space-3);
        }}

        /* Status container */
        [data-testid="stStatusWidget"] {{
            background: white;
            border-radius: var(--radius-md);
            box-shadow: var(--shadow-md);
        }}

        /* ═══════════════════════════════════════════════════════════════════
           ALERTS & MESSAGES
           ═══════════════════════════════════════════════════════════════════ */
        .stAlert {{
            border-radius: var(--radius-md);
            border: none;
            box-shadow: var(--shadow-sm);
        }}

        [data-testid="stAlert"] {{
            padding: var(--space-2) var(--space-3);
        }}

        /* Success message */
        .stSuccess {{
            background: linear-gradient(135deg, #ECFDF5 0%, #D1FAE5 100%);
            border-left: 4px solid {WB_SUCCESS};
        }}

        /* Info message */
        .stInfo {{
            background: linear-gradient(135deg, #EFF6FF 0%, #DBEAFE 100%);
            border-left: 4px solid var(--primary);
        }}

        /* Warning message */
        .stWarning {{
            background: linear-gradient(135deg, #FFFBEB 0%, #FEF3C7 100%);
            border-left: 4px solid #F59E0B;
        }}

        /* Error message */
        .stError {{
            background: linear-gradient(135deg, #FEF2F2 0%, #FEE2E2 100%);
            border-left: 4px solid #EF4444;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           TAGS & BADGES
           ═══════════════════════════════════════════════════════════════════ */
        [data-baseweb="tag"] {{
            background: linear-gradient(135deg, {WB_PRIMARY} 0%, {WB_SECONDARY} 100%);
            border-radius: var(--radius-sm);
            font-weight: 500;
            padding: 4px 12px;
        }}

        [data-baseweb="tag"], [data-baseweb="tag"] * {{
            color: white !important;
        }}

        [data-baseweb="tag"] [data-baseweb="close"] svg {{
            fill: white !important;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           LINKS
           ═══════════════════════════════════════════════════════════════════ */
        a {{
            color: var(--primary) !important;
            text-decoration: none;
            font-weight: 500;
            transition: var(--transition);
        }}

        a:hover {{
            color: var(--accent) !important;
            text-decoration: underline;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           DIVIDERS
           ═══════════════════════════════════════════════════════════════════ */
        hr {{
            border: none;
            height: 1px;
            background: linear-gradient(90deg, transparent, var(--border), transparent);
            margin: var(--space-4) 0;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           PROGRESS BAR
           ═══════════════════════════════════════════════════════════════════ */
        .stProgress > div > div > div {{
            background: linear-gradient(90deg, {WB_PRIMARY} 0%, {WB_ACCENT} 100%);
            border-radius: var(--radius-sm);
        }}

        .stProgress > div > div {{
            background: var(--light);
            border-radius: var(--radius-sm);
        }}

        /* ═══════════════════════════════════════════════════════════════════
           METRIC CARDS
           ═══════════════════════════════════════════════════════════════════ */
        [data-testid="stMetric"] {{
            background: white;
            padding: var(--space-3);
            border-radius: var(--radius-lg);
            border: 1px solid var(--border);
            box-shadow: var(--shadow-sm);
        }}

        [data-testid="stMetric"] label {{
            color: var(--muted) !important;
            font-size: 0.85rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}

        [data-testid="stMetric"] [data-testid="stMetricValue"] {{
            color: var(--text) !important;
            font-weight: 700;
            font-size: 1.8rem;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           PLOTLY CHARTS
           ═══════════════════════════════════════════════════════════════════ */
        [data-testid="stPlotlyChart"] {{
            background: white;
            border-radius: var(--radius-lg);
            box-shadow: var(--shadow-md);
            padding: var(--space-2);
            border: 1px solid var(--border);
        }}

        /* ═══════════════════════════════════════════════════════════════════
           JSON DISPLAY
           ═══════════════════════════════════════════════════════════════════ */
        [data-testid="stJson"] {{
            background: var(--light);
            border-radius: var(--radius-md);
            padding: var(--space-2);
            border: 1px solid var(--border);
        }}

        /* ═══════════════════════════════════════════════════════════════════
           CAPTION TEXT
           ═══════════════════════════════════════════════════════════════════ */
        .stCaption {{
            color: var(--muted) !important;
            font-size: 0.85rem;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           MULTISELECT
           ═══════════════════════════════════════════════════════════════════ */
        [data-baseweb="multi-select"] {{
            border-radius: var(--radius-sm);
        }}

        [data-baseweb="multi-select"] > div {{
            border: 2px solid var(--border) !important;
            border-radius: var(--radius-sm) !important;
            min-height: 42px;
        }}

        [data-baseweb="multi-select"] > div:hover {{
            border-color: var(--primary) !important;
        }}

        /* ═══════════════════════════════════════════════════════════════════
           SLIDER
           ═══════════════════════════════════════════════════════════════════ */
        [data-testid="stSlider"] [data-baseweb="slider"] > div > div {{
            background: var(--light) !important;
        }}

        [data-testid="stSlider"] [role="slider"] {{
            background: var(--primary) !important;
            border: 3px solid white !important;
            box-shadow: var(--shadow-md);
        }}

        /* ═══════════════════════════════════════════════════════════════════
           COLOR PICKER
           ═══════════════════════════════════════════════════════════════════ */
        [data-testid="stColorPicker"] > div {{
            border-radius: var(--radius-sm);
            overflow: hidden;
            box-shadow: var(--shadow-sm);
        }}

    </style>
    """,
    unsafe_allow_html=True,
)

# Compact density adjustments
if density_mode == 'Compact':
    st.markdown(
        """
        <style>
            [data-testid="stAppViewContainer"] {
                padding: var(--space-2);
            }
            .stButton > button {
                padding: 8px 16px;
                margin: var(--space-1) 0;
            }
            .block-container {
                padding: var(--space-2) var(--space-3);
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

# ──────────────────────────────────────────────────────────────────────────────
# CORPORATE HEADER - Light background with dark text for contrast
# ──────────────────────────────────────────────────────────────────────────────
st.markdown(
    f"""
    <div style="
        background: white;
        padding: 32px 40px;
        border-radius: 16px;
        margin-bottom: 32px;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
        border: 1px solid {WB_BORDER};
        border-left: 5px solid {WB_PRIMARY};
    ">
        <div style="display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 20px;">
            <div>
                <div style="display: flex; align-items: center; gap: 16px; margin-bottom: 8px;">
                    <div style="
                        background: {WB_PRIMARY};
                        border-radius: 12px;
                        padding: 12px;
                        display: flex;
                        align-items: center;
                        justify-content: center;
                    ">
                        <svg width="32" height="32" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                            <path d="M3 3v18h18"/>
                            <path d="M18.7 8l-5.1 5.2-2.8-2.7L7 14.3"/>
                        </svg>
                    </div>
                    <h1 style="color: {WB_TEXT} !important; margin: 0; font-size: 2rem; font-weight: 700; letter-spacing: -0.02em;">
                        Vintage Default-Rate Analytics
                    </h1>
                </div>
                <p style="color: {WB_MUTED} !important; margin: 0; font-size: 1rem; max-width: 500px;">
                    Enterprise-grade loan portfolio analysis with integrity validation and performance tracking
                </p>
            </div>
            <div style="display: flex; gap: 24px;">
                <div style="text-align: center;">
                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.1em;">Platform</div>
                    <div style="color: {WB_PRIMARY}; font-weight: 600; font-size: 1rem;">QOB Analysis</div>
                </div>
                <div style="width: 1px; background: {WB_BORDER};"></div>
                <div style="text-align: center;">
                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.1em;">Version</div>
                    <div style="color: {WB_PRIMARY}; font-weight: 600; font-size: 1rem;">2.0</div>
                </div>
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

MAX_MB = 50
RESERVED_COLS = {
    'Loan ID','Origination date','Maturity date','Observation date',
    'Days past due','Origination amount','Current amount',
    'Vintage','MOB','QOB','is_def','is_def_cum'
}

# ──────────────────────────────────────────────────────────────────────────────
# Helpers: schema, typing, ageing
# ──────────────────────────────────────────────────────────────────────────────
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    key = {str(c).strip().lower(): c for c in df.columns}
    def find(*names):
        for n in names:
            c = key.get(n.lower())
            if c is not None:
                return c
        return None
    col_loan = find('loan id','loan_id','loan number','loan no','id')
    col_dpd  = find('days past due','dpd','days_past_due')
    col_orig = find('origination date','origination_date','orig date','orig_date')
    col_obs  = find('observation date','observation_date','obs date','obs_date')
    col_mat  = find('maturity date','maturity_date','mat date','mat_date')
    col_orig_amt = find('origination amount','origination_amount','original amount','principal','orig amt','orig_amt')
    col_curr_amt = find('current amount','current_amount','balance','outstanding','current bal','current_bal')
    missing = []
    if not col_loan:     missing.append('Loan ID')
    if not col_dpd:      missing.append('Days past due')
    if not col_orig:     missing.append('Origination date')
    if not col_obs:      missing.append('Observation date')
    if not col_mat:      missing.append('Maturity date')
    if not col_orig_amt: missing.append('Origination amount')
    if missing:
        raise KeyError(f"Missing required column(s): {', '.join(missing)}")
    rename_map = {
        col_loan:     'Loan ID',
        col_dpd:      'Days past due',
        col_orig:     'Origination date',
        col_obs:      'Observation date',
        col_mat:      'Maturity date',
        col_orig_amt: 'Origination amount',
    }
    if col_curr_amt:
        rename_map[col_curr_amt] = 'Current amount'
    df = df.rename(columns=rename_map)
    if 'Current amount' not in df.columns:
        df['Current amount'] = np.nan
    return df

def ensure_types(df: pd.DataFrame, keep_originals: bool = False) -> pd.DataFrame:
    df = df.copy()
    if keep_originals:
        for c in ['Origination date','Observation date','Maturity date','Days past due',
                  'Origination amount','Current amount']:
            df[f'__orig_{c}'] = df[c]
    for c in ['Origination date','Observation date','Maturity date']:
        if not pd.api.types.is_datetime64_any_dtype(df[c]):
            df[c] = pd.to_datetime(df[c], errors='coerce')
    for c in ['Days past due','Origination amount','Current amount']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    return df

VINTAGE_GRANULARITY_OPTIONS = {
    'Monthly': 'M',
    'Quarterly': 'Q',
    'Semi-annual (6 months)': 'H',
    'Yearly': 'Y',
}

def add_vintage_mob(df: pd.DataFrame, granularity: str = 'Q') -> pd.DataFrame:
    """Add Vintage cohort label and monthly MOB.

    granularity: 'M' (monthly, YYYY-MM), 'Q' (quarterly, YYYYQx — default),
                 'H' (semi-annual, YYYYHx), 'Y' (yearly, YYYY).
    """
    df = df.copy()
    g = (granularity or 'Q').upper()
    od = df['Origination date']
    if g == 'M':
        df['Vintage'] = od.dt.to_period('M').astype(str)
    elif g == 'H':
        half = ((od.dt.month - 1) // 6 + 1).astype('Int64')
        df['Vintage'] = od.dt.year.astype('Int64').astype(str) + 'H' + half.astype(str)
    elif g == 'Y':
        df['Vintage'] = od.dt.year.astype('Int64').astype(str)
    else:  # 'Q' default
        df['Vintage'] = od.dt.to_period('Q').astype(str)
    mob = ((df['Observation date'].dt.year - df['Origination date'].dt.year) * 12
          + (df['Observation date'].dt.month - df['Origination date'].dt.month)) + 1
    df['MOB'] = mob
    df = df[df['MOB'] > 0]
    return df

def add_qob(df: pd.DataFrame) -> pd.DataFrame:
    """Quarter-on-book (QOB) as small int: stable quarterly ageing."""
    df = df.copy()
    oy = df['Origination date'].dt.year
    oq = df['Origination date'].dt.quarter
    sy = df['Observation date'].dt.year
    sq = df['Observation date'].dt.quarter
    o_code = oy * 4 + (oq - 1)
    s_code = sy * 4 + (sq - 1)
    qob = (s_code - o_code + 1)
    df['QOB'] = pd.to_numeric(qob, errors='coerce')
    df = df[df['QOB'].gt(0).fillna(False)]
    df['QOB'] = df['QOB'].astype('int16', copy=False)
    return df

# ──────────────────────────────────────────────────────────────────────────────
# Caching
# ──────────────────────────────────────────────────────────────────────────────
def _df_key(df: pd.DataFrame) -> str:
    cols = [c for c in ['Loan ID','Origination date','Observation date','Maturity date',
                        'Days past due','Origination amount','Current amount'] if c in df.columns]
    if not cols:
        cols = list(df.columns)
    parts = [df[cols].head(100000), df[cols].tail(1000)]
    sample = pd.concat(parts).drop_duplicates()
    h = pd.util.hash_pandas_object(sample, index=True).values
    m = hashlib.blake2b(digest_size=16)
    m.update(h.tobytes()); m.update(str(df.shape).encode()); m.update(",".join(map(str, sample.columns)).encode())
    return m.hexdigest()

@cache_data_smart(show_spinner=False, max_entries=2, hash_funcs={pd.DataFrame: _df_key})
def prepare_base_cached(raw_df: pd.DataFrame, vintage_granularity: str = 'Q') -> pd.DataFrame:
    logger.info("prepare_base_cached: %d rows input", len(raw_df))
    dfn = normalize_columns(raw_df)
    dfn = ensure_types(dfn)
    dfn = add_vintage_mob(dfn, granularity=vintage_granularity)
    dfn['Vintage'] = dfn['Vintage'].astype('category')
    dfn['Loan ID'] = dfn['Loan ID'].astype('category')
    for c in ['Days past due','Origination amount','Current amount']:
        if c in dfn.columns:
            dfn[c] = dfn[c].astype('float32')
    dfn = dfn.drop_duplicates(subset=['Loan ID','Observation date'], keep='last')
    dfn = dfn.sort_values(['Loan ID','Observation date'], kind='mergesort')
    _log_resource_usage("after prepare_base_cached")
    gc.collect()
    return dfn

@cache_data_smart(show_spinner=False, max_entries=1)
def load_full(file_bytes: bytes, sheet: str, header: int):
    bio = BytesIO(file_bytes)
    return pd.read_excel(bio, sheet_name=sheet, header=header, engine='openpyxl')

# ──────────────────────────────────────────────────────────────────────────────
# Fast per-loan cumulative OR (Numba or pandas fallback)
# ──────────────────────────────────────────────────────────────────────────────
if NUMBA_OK:
    from numba import njit
    @njit(cache=True, fastmath=True)
    def _cum_or_by_group(codes: np.ndarray, flags: np.ndarray) -> np.ndarray:
        n = flags.size
        out = np.empty(n, np.uint8)
        if n == 0: return out
        prev = codes[0]; acc = flags[0] != 0; out[0] = 1 if acc else 0
        for i in range(1, n):
            c = codes[i]
            if c != prev:
                prev = c; acc = flags[i] != 0
            else:
                if flags[i] != 0: acc = True
            out[i] = 1 if acc else 0
        return out
else:
    def _cum_or_by_group(codes: np.ndarray, flags: np.ndarray) -> np.ndarray:
        s_codes = pd.Series(codes, copy=False)
        s_flags = pd.Series(flags, copy=False)
        return (s_flags.groupby(s_codes, sort=False).cummax().astype(np.uint8).to_numpy(copy=False))

# ──────────────────────────────────────────────────────────────────────────────
# Progress helper
# ──────────────────────────────────────────────────────────────────────────────
def mk_progress_updater(bar, steps: int = 5) -> Callable[[str], None]:
    ctr = {"i": 0, "steps": max(1, int(steps))}
    def _update(msg: str):
        ctr["i"] += 1
        bar.progress(min(ctr["i"]/ctr["steps"], 1.0), text=msg)
    return _update

# ──────────────────────────────────────────────────────────────────────────────
# Chart pipeline (supports QOB and MOB, smoothing toggles, cure-adjusted mode)
# ──────────────────────────────────────────────────────────────────────────────
def build_chart_data_fast(raw_df: pd.DataFrame, dpd_threshold: int,
                          max_periods: int = 20,
                          granularity: str = "QOB",
                          smooth: bool = True,
                          force_monotone: bool = True,
                          cure_adjusted: bool = False,
                          exclude_indices: Optional[set] = None,
                          pd_by_amount: bool = False,
                          vintage_granularity: str = 'Q',
                          prog: Optional[Callable[[str], None]] = None) -> tuple[pd.DataFrame, dict]:
    if prog: prog("Preparing base dataset …")
    base = prepare_base_cached(raw_df, vintage_granularity=vintage_granularity)

    if exclude_indices:
        base = base.loc[~base.index.isin(exclude_indices)]

    period_col = granularity.upper()
    if period_col == "QOB":
        base = add_qob(base)
    else:
        period_col = "MOB"

    if prog: prog("Computing default flags …")
    flags = (base['Days past due'].to_numpy(np.float32, copy=False) >= dpd_threshold).astype(np.uint8, copy=False)

    if cure_adjusted:
        is_def = flags
    else:
        codes = base['Loan ID'].cat.codes.to_numpy(np.int64, copy=False)
        if prog: prog("Applying per-loan cummax …")
        is_def = _cum_or_by_group(codes, flags)

    if prog: prog(f"Aggregating cohorts ({period_col}) …")
    work_cols = {
        'Vintage': base['Vintage'].to_numpy(),
        period_col: base[period_col].to_numpy(),
        'is_def': is_def,
        'LoanID': base['Loan ID'].to_numpy(),
    }
    if pd_by_amount:
        work_cols['Origination_amount'] = base['Origination amount'].to_numpy()
        work_cols['Current_amount'] = base['Current amount'].to_numpy()
        if not cure_adjusted:
            # Raw flags needed to find the actual first-default observation
            raw_flags = (base['Days past due'].to_numpy(np.float32, copy=False)
                         >= dpd_threshold).astype(np.uint8, copy=False)
            work_cols['raw_def'] = raw_flags
    work = pd.DataFrame(work_cols)

    if work.empty:
        if prog: prog("No data to plot.")
        return pd.DataFrame(), {}

    if pd_by_amount:
        # Denominator: total origination amount per vintage (one entry per loan)
        loan_first = work.drop_duplicates(subset='LoanID', keep='first')
        vintage_total_orig_amt = loan_first.groupby('Vintage')['Origination_amount'].sum()
        cohort_sizes = vintage_total_orig_amt.to_dict()
    else:
        # Fixed denominator: total unique loans per vintage (matches table logic)
        vintage_total_loans = work.groupby('Vintage')['LoanID'].nunique()
        cohort_sizes = vintage_total_loans.to_dict()

    all_periods = pd.to_numeric(work[period_col], errors='coerce').dropna()
    max_p = min(int(all_periods.max()), max_periods) if len(all_periods) > 0 else max_periods

    if pd_by_amount:
        # ── Amount-weighted PD ──
        if cure_adjusted:
            # Sum Current Amount of currently-defaulted loans per (vintage, period)
            def_rows = work.loc[work['is_def'] == 1]
            # Deduplicate per (vintage, period, loan): take last observation's amount
            loan_period_amt = (def_rows.groupby(['Vintage', period_col, 'LoanID'], sort=False)
                               ['Current_amount'].last().reset_index())
            agg = (loan_period_amt.groupby(['Vintage', period_col], sort=False)
                   ['Current_amount'].sum().reset_index(name='total_default_amount'))
            agg['total_orig'] = agg['Vintage'].map(vintage_total_orig_amt)
            agg['default_rate'] = (agg['total_default_amount'] / agg['total_orig']).astype('float32')
            wide = (agg.pivot(index=period_col, columns='Vintage', values='default_rate')
                       .sort_index()
                       .reindex(range(1, max_p + 1)))
        else:
            # Non-cure-adjusted: find first actual default observation per loan,
            # get Current Amount at that observation, cumulate over periods
            actual_defaults = work.loc[work['raw_def'] == 1]
            first_def = (actual_defaults.sort_values(period_col)
                         .drop_duplicates(subset='LoanID', keep='first'))
            # Sum new default amounts at each period per vintage
            new_def_amounts = (first_def.groupby(['Vintage', period_col])['Current_amount']
                               .sum().reset_index(name='new_default_amount'))
            # Build full (vintage, period) grid and compute cumulative
            vintages = work['Vintage'].unique()
            periods = list(range(1, max_p + 1))
            full_idx = pd.MultiIndex.from_product([vintages, periods],
                                                  names=['Vintage', period_col])
            cum_df = (new_def_amounts.set_index(['Vintage', period_col])
                                     .reindex(full_idx, fill_value=0)
                                     .reset_index())
            cum_df['cum_default_amount'] = cum_df.groupby('Vintage')['new_default_amount'].cumsum()
            cum_df['total_orig'] = cum_df['Vintage'].map(vintage_total_orig_amt)
            cum_df['default_rate'] = (cum_df['cum_default_amount'] / cum_df['total_orig']).astype('float32')
            wide = (cum_df.pivot(index=period_col, columns='Vintage', values='default_rate')
                          .sort_index()
                          .reindex(range(1, max_p + 1)))
    else:
        # ── Loan-count PD (original logic) ──
        # Deduplicate to one row per (loan, period): take max of is_def
        loan_period = (work.groupby(['Vintage', period_col, 'LoanID'], sort=False)['is_def']
                           .max().reset_index())
        loan_period[period_col] = pd.to_numeric(loan_period[period_col], errors='coerce').astype('int16')

        if cure_adjusted:
            # Cure-adjusted: count currently-defaulted unique loans per (vintage, period)
            agg = (loan_period.groupby(['Vintage', period_col], sort=False)
                   .agg(total_default=('is_def', 'sum'))
                   .reset_index())
            agg['total_loans'] = agg['Vintage'].map(vintage_total_loans)
            agg['default_rate'] = (agg['total_default'] / agg['total_loans']).astype('float32')
            wide = (agg.pivot(index=period_col, columns='Vintage', values='default_rate')
                       .sort_index()
                       .reindex(range(1, max_p + 1)))
        else:
            # Non-cure-adjusted: cumulative unique defaults by period
            # Find the earliest period each loan first defaults
            defaulted = loan_period.loc[loan_period['is_def'] == 1]
            first_def = (defaulted.groupby(['Vintage', 'LoanID'])[period_col]
                                  .min().reset_index()
                                  .rename(columns={period_col: 'first_def_period'}))

            # Count new defaults at each period per vintage
            new_defaults = (first_def.groupby(['Vintage', 'first_def_period'])
                                     .size().reset_index(name='new_defaults')
                                     .rename(columns={'first_def_period': period_col}))

            # Build full (vintage, period) grid and compute cumulative defaults
            vintages = work['Vintage'].unique()
            periods = list(range(1, max_p + 1))
            full_idx = pd.MultiIndex.from_product([vintages, periods],
                                                  names=['Vintage', period_col])
            cum_df = (new_defaults.set_index(['Vintage', period_col])
                                  .reindex(full_idx, fill_value=0)
                                  .reset_index())
            cum_df['cum_defaults'] = cum_df.groupby('Vintage')['new_defaults'].cumsum()
            cum_df['total_loans'] = cum_df['Vintage'].map(vintage_total_loans)
            cum_df['default_rate'] = (cum_df['cum_defaults'] / cum_df['total_loans']).astype('float32')

            wide = (cum_df.pivot(index=period_col, columns='Vintage', values='default_rate')
                          .sort_index()
                          .reindex(range(1, max_p + 1)))

    # Save raw endpoint values before smoothing (for anchoring)
    raw_endpoints = {}
    for col in wide.columns:
        s = wide[col].dropna()
        if len(s) > 0:
            raw_endpoints[col] = s.iloc[-1]

    if smooth:
        wide = wide.rolling(2, 1, center=True).mean()
    if force_monotone:
        wide = wide.cummax()

    # Anchor: restore the last valid data point to raw value so chart
    # endpoints match the table Cum PD exactly (only for non-cure-adjusted,
    # since cure-adjusted intentionally shows current status, not cumulative)
    if not cure_adjusted and (smooth or force_monotone):
        for col, raw_val in raw_endpoints.items():
            s = wide[col].dropna()
            if len(s) > 0:
                wide.loc[s.index[-1], col] = raw_val

    wide = wide.astype('float32')
    wide.index.name = period_col
    return wide, cohort_sizes

# ──────────────────────────────────────────────────────────────────────────────
# Plotting: % y-axis, months x-axis (QOB*3), optional legend
# ──────────────────────────────────────────────────────────────────────────────

def plot_curves_percent_with_months(df_wide: pd.DataFrame,
                                    title: str,
                                    show_legend: bool = True,
                                    legend_limit: int = 40,
                                    palette: str = "Gradient",
                                    base_color: Optional[str] = None,
                                    line_width: int = 1,
                                    cohort_sizes: Optional[dict] = None,
                                    line_style: str = "solid",
                                    show_markers: bool = False,
                                    chart_height: int = 550,
                                    title_font_size: int = 18,
                                    axis_font_size: int = 13,
                                    show_grid: bool = True,
                                    y_axis_max: float = 0.0,
                                    bg_color: str = "#FFFFFF",
                                    pd_by_amount: bool = False):
    if df_wide.empty:
        st.info('Not enough data to plot.')
        return None

    idx = df_wide.index.to_numpy()
    name = (df_wide.index.name or "").upper()
    x_months = idx * 3 if name == "QOB" else idx  # MOB already months

    Y = df_wide.to_numpy(dtype='float32')
    M, N = Y.shape
    target_points = 200_000
    step = max(1, int(np.ceil((M * N) / target_points)))
    if step > 1:
        x_months = x_months[::step]
        Y = Y[::step, :]

    def _generate_palette(base: str, n: int) -> list[str]:
        r, g, b = hex_to_rgb(base)
        palette = []
        for i in range(n):
            ratio = 0.2 + 0.8 * (i / max(n - 1, 1))
            ri = int(r + (255 - r) * ratio)
            gi = int(g + (255 - g) * ratio)
            bi = int(b + (255 - b) * ratio)
            palette.append(f'rgb({ri},{gi},{bi})')
        return palette

    palette_choice = palette.lower()
    base_color = base_color or st.get_option("theme.primaryColor") or "#1f77b4"
    if palette_choice == "gradient":
        palette_colors = _generate_palette(base_color, N)
    elif palette_choice == "plotly":
        palette_colors = qualitative.Plotly
    elif palette_choice == "viridis":
        palette_colors = sequential.Viridis
    else:
        palette_colors = qualitative.Plotly
    if len(palette_colors) < N:
        times = int(np.ceil(N / len(palette_colors)))
        palette_colors = (palette_colors * times)[:N]
    else:
        palette_colors = palette_colors[:N]

    trace_mode = 'lines+markers' if show_markers else 'lines'
    marker_cfg = dict(size=4) if show_markers else None

    fig = go.Figure()
    show_leg = show_legend and (df_wide.shape[1] <= legend_limit)
    for i, col in enumerate(df_wide.columns):
        label = str(col)
        if cohort_sizes and col in cohort_sizes:
            if pd_by_amount:
                label = f"{col} (${cohort_sizes[col]:,.0f})"
            else:
                label = f"{col} (n={int(cohort_sizes[col]):,})"
        trace_kwargs = dict(
            x=x_months,
            y=Y[:, i],
            mode=trace_mode,
            name=label,
            line=dict(color=palette_colors[i], width=line_width, dash=line_style),
            hovertemplate=f"Vintage: {col}<br>Month: %{{x}}<br>Default rate: %{{y:.2%}}<extra></extra>"
        )
        if marker_cfg:
            trace_kwargs['marker'] = marker_cfg
        fig.add_trace(go.Scatter(**trace_kwargs))

    grid_color = '#E2E8F0' if show_grid else 'rgba(0,0,0,0)'

    yaxis_cfg = dict(
        title=dict(text='Cumulative Default Rate (by Amount)' if pd_by_amount else 'Cumulative Default Rate',
                   font=dict(size=axis_font_size, color='#64748B')),
        tickformat='.2%',
        tickfont=dict(size=11, color='#64748B'),
        gridcolor=grid_color,
        linecolor='#E2E8F0',
        zeroline=False,
        showgrid=show_grid,
    )
    if y_axis_max > 0:
        yaxis_cfg['range'] = [0, y_axis_max / 100.0]

    fig.update_layout(
        title=dict(
            text=title,
            font=dict(size=title_font_size, color='#002244', family='Inter, sans-serif'),
            x=0.5,
            xanchor='center'
        ),
        xaxis=dict(
            title=dict(text='Deal Age (months)', font=dict(size=axis_font_size, color='#64748B')),
            tickfont=dict(size=11, color='#64748B'),
            gridcolor=grid_color,
            linecolor='#E2E8F0',
            zeroline=False,
            showgrid=show_grid,
        ),
        yaxis=yaxis_cfg,
        hovermode='x unified',
        hoverlabel=dict(
            bgcolor='white',
            bordercolor='#E2E8F0',
            font=dict(size=12, color='#002244', family='Inter, sans-serif')
        ),
        showlegend=show_leg,
        legend=dict(
            bgcolor='rgba(255,255,255,0.95)',
            bordercolor='#E2E8F0',
            borderwidth=1,
            font=dict(color='#002244', size=11, family='Inter, sans-serif'),
            title=dict(text='Vintage Cohorts', font=dict(size=12, color='#64748B')),
        ),
        height=chart_height,
        plot_bgcolor=bg_color,
        paper_bgcolor='white',
        margin=dict(l=60, r=30, t=60, b=50),
        font=dict(family='Inter, sans-serif'),
    )

    return fig
# ──────────────────────────────────────────────────────────────────────────────
# Integrity checks (vectorized)
# ──────────────────────────────────────────────────────────────────────────────
def _years_list(ser: pd.Series) -> list:
    try:
        years = pd.to_datetime(ser, errors='coerce').dt.year.dropna().astype(int)
        return sorted(np.unique(years).tolist())
    except Exception:
        return []

def run_integrity_checks(df: pd.DataFrame, dpd_threshold: int, gap_days: int = 120, after_mat_tol_days: int = 31,
                         vintage_granularity: str = 'Q'):
    summary = {}
    vintage_issues = []
    row_issue_map: dict[int, list[str]] = {}

    def track(mask: pd.Series, issue: str, data: pd.DataFrame = None):
        """Add up to 500 offending rows for `issue` using the index of `data`."""
        nonlocal row_issue_map
        if data is None:
            data = dfn
        mask = mask.fillna(False)
        if mask.any():
            idxs = data[mask].head(500).index
            for i in idxs:
                row_issue_map.setdefault(int(i), []).append(issue)

    try:
        dfn = normalize_columns(df)
    except KeyError as e:
        return {'fatal': str(e)}, pd.DataFrame(), pd.DataFrame()
    dfn = ensure_types(dfn, keep_originals=True)

    # Nulls + non-parsables
    for c in ['Loan ID','Origination date','Observation date','Maturity date',
              'Days past due','Origination amount','Current amount']:
        summary[f'Nulls in {c}'] = int(dfn[c].isna().sum())
    for c in ['Origination date','Observation date','Maturity date',
              'Days past due','Origination amount','Current amount']:
        coerced_nan = dfn[c].isna() & dfn[f'__orig_{c}'].notna()
        summary[f'Non-parsable {c} (after coercion)'] = int(coerced_nan.sum())
        track(coerced_nan, f'Non-parsable {c}')


    summary['Years (Origination)'] = _years_list(dfn['Origination date'])
    summary['Years (Observation)'] = _years_list(dfn['Observation date'])
    summary['Years (Maturity)']    = _years_list(dfn['Maturity date'])


        # Date logic
    mask_obs_before_orig = dfn['Observation date'] < dfn['Origination date']
    summary['Observation before Origination'] = int(mask_obs_before_orig.sum())
    track(mask_obs_before_orig, 'Observation before Origination')

    mask_mat_before_orig = dfn['Maturity date'] < dfn['Origination date']
    summary['Maturity before Origination'] = int(mask_mat_before_orig.sum())
    track(mask_mat_before_orig, 'Maturity before Origination')

    mask_obs_after_mat = dfn['Observation date'] > (dfn['Maturity date'] + pd.to_timedelta(after_mat_tol_days, unit='D'))
    summary['Observation well after Maturity (> tol)'] = int(mask_obs_after_mat.sum())
    track(mask_obs_after_mat, 'Observation well after Maturity')


    

    # Snapshot uniqueness & continuity
    dup_mask = dfn.duplicated(subset=['Loan ID', 'Observation date'], keep=False)
    summary['Duplicate snapshots (Loan ID + Observation date)'] = int(dup_mask.sum())
    track(dup_mask, 'Duplicate snapshot')

    multi_orig = dfn.groupby('Loan ID')['Origination date'].nunique() > 1
    summary['Loans with multiple Origination dates'] = int(multi_orig.sum())
    if multi_orig.any():
        mask_multi_orig = dfn['Loan ID'].isin(multi_orig[multi_orig].index)
        track(mask_multi_orig, 'Multiple origination dates')

    multi_mat = dfn.groupby('Loan ID')['Maturity date'].nunique() > 1
    summary['Loans with changing Maturity date'] = int(multi_mat.sum())
    if multi_mat.any():
        mask_multi_mat = dfn['Loan ID'].isin(multi_mat[multi_mat].index)
        track(mask_multi_mat, 'Maturity date changed')

    dfn = dfn.sort_values(['Loan ID','Observation date'])
    diffs_days = dfn.groupby('Loan ID', sort=False)['Observation date'].diff().dt.days
    out_of_order = diffs_days < 0
    large_gap = diffs_days > gap_days
    summary['Out-of-order snapshots'] = int(out_of_order.fillna(False).sum())
    summary[f'Large gaps in Observation (>{gap_days} days)'] = int(large_gap.fillna(False).sum())
    track(out_of_order, 'Out-of-order snapshots')
    track(large_gap, 'Large gap between snapshots')

    # DPD quality
    neg_dpd_mask = dfn['Days past due'] < 0
    summary['Negative Days past due'] = int(neg_dpd_mask.sum())
    track(neg_dpd_mask, 'Negative DPD')

    non_int_mask = dfn['Days past due'].notna() & ((dfn['Days past due'] % 1) != 0)
    summary['Non-integer DPD values'] = int(non_int_mask.sum())
    track(non_int_mask, 'Non-integer DPD')

    extreme_dpd_mask = dfn['Days past due'] > 3650
    summary['Extreme DPD (> 3650)'] = int(extreme_dpd_mask.sum())
    track(extreme_dpd_mask, 'Extreme DPD')

    prev_dpd = dfn.groupby('Loan ID', sort=False)['Days past due'].shift()
    cure_mask = (prev_dpd >= 180) & (dfn['Days past due'] == 0)
    summary['Sudden cures (>=180 to 0 next)'] = int(cure_mask.fillna(False).sum())
    track(cure_mask, 'Sudden cure 180->0')
   
    # Amounts
    orig_amt_nonpos = dfn['Origination amount'] <= 0
    summary['Origination amount <= 0'] = int(orig_amt_nonpos.sum())
    track(orig_amt_nonpos, 'Non-positive Origination amount')

    orig_amt_changes = dfn.groupby('Loan ID')['Origination amount'].nunique() > 1
    summary['Loans with changing Origination amount'] = int(orig_amt_changes.sum())
    if orig_amt_changes.any():
        mask_orig_amt_changes = dfn['Loan ID'].isin(orig_amt_changes[orig_amt_changes].index)
        track(mask_orig_amt_changes, 'Origination amount changed')

    curr_amt_neg = dfn['Current amount'] < 0
    summary['Negative Current amount'] = int(curr_amt_neg.sum())
    track(curr_amt_neg, 'Negative Current amount')

    curr_gt_orig = dfn['Current amount'] > dfn['Origination amount']
    summary['Current amount > Origination amount'] = int(curr_gt_orig.sum())
    track(curr_gt_orig, 'Current > Origination')

    # Consistency + QOB aggregation sanity
    dfd = add_vintage_mob(dfn, granularity=vintage_granularity).sort_values(['Loan ID','Observation date'])
    dfd['is_def'] = (dfd['Days past due'] >= dpd_threshold).astype(np.uint8)
    dfd['is_def_cum'] = dfd.groupby('Loan ID', sort=False)['is_def'].cummax()
    def_cum_reset = dfd.groupby('Loan ID', sort=False)['is_def_cum'].diff() < 0
    summary['is_def_cum resets (should be 0)'] = int(def_cum_reset.fillna(False).sum())
    track(def_cum_reset, 'is_def_cum reset', data=dfd)

    dfd_q = add_qob(dfd)
    agg = dfd_q.groupby(['Vintage','QOB'], sort=False).agg(
        total_loans=('Loan ID','nunique'),
        total_default=('is_def_cum','sum')
    ).reset_index()
    vintages = sorted(agg['Vintage'].unique().tolist())
    summary['Vintages observed'] = vintages

    incr_rows = []
    for v in vintages:
        sub = agg[agg['Vintage']==v].sort_values('QOB')
        t = sub['total_loans'].to_numpy()
        if len(t) >= 2:
            inc = (np.diff(t) > 0)
            if inc.any():
                where = np.where(inc)[0]
                for idx in where:
                    q_prev = int(sub.iloc[idx]['QOB'])
                    q_curr = int(sub.iloc[idx+1]['QOB'])
                    incr_rows.append({'Vintage': v, 'QOB_prev': q_prev, 'QOB_curr': q_curr,
                                      'total_prev': int(sub.iloc[idx]['total_loans']),
                                      'total_curr': int(sub.iloc[idx+1]['total_loans']),
                                      'issue': 'Denominator increased with QOB'})
    summary['Vintage denominators increasing (count, QOB)'] = len(incr_rows)
    if incr_rows:
        vintage_issues.append(pd.DataFrame(incr_rows))

    qob1 = agg[agg['QOB']==1].set_index('Vintage')['total_loans']
    orig_counts = dfd_q.groupby('Vintage')['Loan ID'].nunique()
    coverage = pd.DataFrame({'orig_loans': orig_counts}).join(qob1.rename('qob1_loans'), how='left')
    coverage['qob1_coverage_%'] = 100 * (coverage['qob1_loans'] / coverage['orig_loans'])
    low_cov = coverage[coverage['qob1_coverage_%'].fillna(0) < 80]
    summary['Vintages with QOB1 coverage < 80%'] = int(len(low_cov))
    if not low_cov.empty:
        tmp = low_cov.reset_index()
        tmp['issue'] = 'Low QOB1 coverage'
        vintage_issues.append(tmp[['Vintage','orig_loans','qob1_loans','qob1_coverage_%','issue']])

    raw = dfd_q.groupby(['Vintage','QOB'], sort=False).agg(default_rate=('is_def_cum','mean')).reset_index()
    non_monotone = []
    for v in vintages:
        sub = raw[raw['Vintage']==v].sort_values('QOB')
        r = sub['default_rate'].to_numpy()
        if len(r) >= 2 and np.any(np.diff(r) < -1e-12):
            non_monotone.append(v)
    summary['Vintages with non-monotone raw default rate (QOB)'] = len(non_monotone)
    if non_monotone:
        vintage_issues.append(pd.DataFrame({'Vintage': non_monotone,'issue': 'Raw default rate not monotone (QOB)'}))

    # Identify loans that disappear between consecutive QOBs for non-monotone vintages
    disappeared_rows = []
    if non_monotone:
        for v in non_monotone:
            v_data = dfd_q[dfd_q['Vintage'] == v]
            qobs = sorted(v_data['QOB'].unique())
            for i in range(len(qobs) - 1):
                q_prev, q_curr = qobs[i], qobs[i + 1]
                ids_prev = set(v_data.loc[v_data['QOB'] == q_prev, 'Loan ID'])
                ids_curr = set(v_data.loc[v_data['QOB'] == q_curr, 'Loan ID'])
                gone = ids_prev - ids_curr
                if gone:
                    for lid in gone:
                        loan_rows = v_data[v_data['Loan ID'] == lid]
                        last_row = loan_rows.sort_values('Observation date').iloc[-1]
                        disappeared_rows.append({
                            'Vintage': v,
                            'Loan ID': lid,
                            'Last seen QOB': int(q_prev),
                            'Missing from QOB': int(q_curr),
                            'Last DPD': last_row.get('Days past due', ''),
                            'Last Observation date': last_row.get('Observation date', ''),
                            'Origination date': last_row.get('Origination date', ''),
                            'Origination amount': last_row.get('Origination amount', ''),
                        })
    disappeared_df = pd.DataFrame(disappeared_rows) if disappeared_rows else pd.DataFrame()
    if row_issue_map:
        sample_indices = list(row_issue_map.keys())
        rows = dfn.loc[sample_indices].copy()
        rows['issues'] = ['; '.join(row_issue_map[idx]) for idx in sample_indices]
        issues_df = rows.reset_index()
    else:
        issues_df = pd.DataFrame()

    vintage_issues_df = pd.concat(vintage_issues, ignore_index=True) if vintage_issues else pd.DataFrame()

    

    summary['Rows (total)'] = int(len(dfn))
    summary['Distinct loans'] = int(dfn['Loan ID'].nunique())
    try:
        summary['Date range (Origination)'] = f"{dfn['Origination date'].min().date()} → {dfn['Origination date'].max().date()}"
        summary['Date range (Observation)'] = f"{dfn['Observation date'].min().date()} → {dfn['Observation date'].max().date()}"
        summary['Date range (Maturity)']    = f"{dfn['Maturity date'].min().date()} → {dfn['Maturity date'].max().date()}"
    except Exception:
        pass

    # Free intermediate DataFrames
    del dfn, dfd, dfd_q
    gc.collect()

    return summary, issues_df, vintage_issues_df, disappeared_df

# Human-readable descriptions for integrity checks
CHECK_DESCRIPTIONS = {
    'Observation before Origination': 'Observation date occurs before origination date for a loan.',
    'Maturity before Origination': 'Maturity date occurs before origination date.',
    'Observation well after Maturity (> tol)': 'Observation date is far beyond maturity date.',
    'Duplicate snapshots (Loan ID + Observation date)': 'Same loan has multiple rows with identical observation date.',
    'Loans with multiple Origination dates': 'A loan appears with more than one origination date.',
    'Loans with changing Maturity date': 'A loan shows different maturity dates across records.',
    'Out-of-order snapshots': 'Observation dates are not in chronological order for a loan.',
    'Large gaps in Observation': 'Long intervals between successive observations for a loan.',
    'Negative Days past due': 'Days past due value is negative.',
    'Non-integer DPD values': 'Days past due is not a whole number.',
    'Extreme DPD (> 3650)': 'Days past due exceeds ten years.',
    'Sudden cures (>=180 to 0 next)': 'DPD drops from ≥180 to 0 in the next snapshot.',
    'Origination amount <= 0': 'Origination amount is non-positive.',
    'Loans with changing Origination amount': 'Origination amount varies for the same loan.',
    'Negative Current amount': 'Current amount is negative.',
    'Current amount > Origination amount': 'Current amount exceeds the origination amount.',
    'is_def_cum resets (should be 0)': 'Cumulative default flag decreases, which should not happen.',
    'Vintage denominators increasing (count, QOB)': 'Number of loans grows with ageing, implying duplicates.',
    'Vintages with QOB1 coverage < 80%': 'Vintages where fewer than 80% of loans have an initial snapshot.',
    'Vintages with non-monotone raw default rate (QOB)': 'Default rate decreases between QOBs for a vintage.',
    'Rows (total)': 'Total rows in dataset.',
    'Distinct loans': 'Count of unique Loan ID values.',
    'Date range (Origination)': 'Earliest and latest origination dates.',
    'Date range (Observation)': 'Earliest and latest observation dates.',
    'Date range (Maturity)': 'Earliest and latest maturity dates.',
    'Years (Origination)': 'Distinct origination years present.',
    'Years (Observation)': 'Distinct observation years present.',
    'Years (Maturity)': 'Distinct maturity years present.',
    'Vintages observed': 'Vintages represented after processing.'
}

def explain_check(name: str) -> str:
    if name.startswith('Nulls in '):
        col = name[len('Nulls in '):]
        return f'Rows where {col} is missing.'
    if name.startswith('Non-parsable '):
        col = name[len('Non-parsable '):].split(' (')[0]
        return f'Values in {col} could not be parsed to the required type.'
    if name.startswith('Large gaps in Observation'):
        return CHECK_DESCRIPTIONS['Large gaps in Observation']
    return CHECK_DESCRIPTIONS.get(name, '')

# ──────────────────────────────────────────────────────────────────────────────
# Exports
# ──────────────────────────────────────────────────────────────────────────────
VINTAGE_ISSUE_EXPLANATIONS = {
    'Denominator increased with QOB': (
        "The loan count within this vintage increased over successive quarters-on-book. "
        "By definition, a closed cohort cannot grow — loans may prepay, mature, or default, "
        "but new originations should never be added retrospectively. An increasing denominator "
        "typically indicates duplicate records, incorrect vintage assignment, or data ingestion "
        "errors. This anomaly directly impacts the reliability of default-rate calculations, "
        "as an inflated denominator understates observed PD."
    ),
    'Low QOB1 coverage': (
        "Fewer than 80% of loans originated in this vintage are represented in the first "
        "quarter-on-book snapshot. Adequate QOB1 coverage is essential for establishing an "
        "accurate cohort baseline. Low coverage may stem from delayed data capture, "
        "incomplete source feeds, or misalignment between origination and reporting periods. "
        "Incomplete initial observation undermines the denominator and may introduce "
        "survivorship bias into subsequent default-rate estimates."
    ),
    'Raw default rate not monotone (QOB)': (
        "The cumulative default rate for this vintage decreases between consecutive quarters-on-book. "
        "Cumulative default rates are monotonically non-decreasing by construction — once a loan "
        "enters default status, it remains classified as such in perpetuity. A declining rate "
        "signals potential data quality issues such as retroactive reclassification of defaults, "
        "inconsistent default flag logic, or denominator instability across observation periods. "
        "This must be investigated before the vintage curve can be used for benchmarking or model calibration."
    ),
}

def export_integrity_pdf(summary: dict, dataset_label: str = 'Full dataset',
                         vintage_issues_df: pd.DataFrame = None) -> bytes:
    import datetime

    buf = BytesIO()
    LINE_H = 0.024
    LINE_H_SM = 0.021
    MARGIN_TOP = 0.91
    MARGIN_BOT = 0.06
    MARGIN_L = 0.07
    MARGIN_R = 0.93
    PAGE_W = 8.27
    PAGE_H = 11.69

    # Corporate colour palette
    C_PRIMARY = '#1E3A8A'
    C_DARK = '#0F172A'
    C_TEXT = '#1E293B'
    C_MUTED = '#64748B'
    C_ACCENT = '#2563EB'
    C_RULE = '#CBD5E1'
    C_RED = '#991B1B'
    C_GREEN = '#047857'
    C_BG_LIGHT = '#F1F5F9'

    def _new_page(pdf_pages):
        fig = plt.figure(figsize=(PAGE_W, PAGE_H))
        ax = fig.add_axes([0, 0, 1, 1])
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis('off')
        return fig, ax, MARGIN_TOP

    def _save_page(pdf_pages, fig):
        pdf_pages.savefig(fig)
        plt.close(fig)

    def _draw_rule(ax, y, color=C_RULE, lw=0.5):
        ax.plot([MARGIN_L, MARGIN_R], [y, y], color=color, lw=lw,
                transform=ax.transAxes, clip_on=False)

    def _need_space(pdf_pages, fig, ax, y, needed, section_name=None):
        if y - needed < MARGIN_BOT:
            _save_page(pdf_pages, fig)
            fig, ax, y = _new_page(pdf_pages)
            if section_name:
                ax.text(0.5, 0.96, 'Data Integrity Report', ha='center',
                        va='top', fontsize=13, weight='bold', color=C_PRIMARY,
                        fontstyle='italic')
                _draw_rule(ax, 0.945, color=C_PRIMARY, lw=0.8)
                y = MARGIN_TOP
        return fig, ax, y

    def _section_header(ax, y, title):
        _draw_rule(ax, y + 0.008, color=C_PRIMARY, lw=1.0)
        ax.text(MARGIN_L, y - 0.005, title.upper(), ha='left', va='top',
                fontsize=12, weight='bold', color=C_PRIMARY,
                fontfamily='sans-serif')
        return y - 0.035

    with PdfPages(buf) as pdf:
        # ── PAGE 1: Cover / Title ──────────────────────────────────────
        fig, ax, y = _new_page(pdf)

        # Top accent bar
        from matplotlib.patches import FancyBboxPatch
        ax.add_patch(FancyBboxPatch((0, 0.88), 1, 0.12,
                     boxstyle="square,pad=0", facecolor=C_PRIMARY, edgecolor='none'))

        ax.text(0.5, 0.955, 'DATA INTEGRITY REPORT', ha='center', va='top',
                fontsize=24, weight='bold', color='white',
                fontfamily='sans-serif')
        ax.text(0.5, 0.915, 'Vintage Default-Rate Analytics', ha='center',
                va='top', fontsize=13, color='#93C5FD', fontfamily='sans-serif')

        y = 0.85

        # Dataset & timestamp line
        ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        ax.text(MARGIN_L, y, f'Dataset:  {dataset_label}', ha='left',
                va='top', fontsize=10.5, color=C_TEXT, weight='bold')
        ax.text(MARGIN_R, y, f'Generated:  {ts}', ha='right',
                va='top', fontsize=10.5, color=C_MUTED)
        y -= 0.035
        _draw_rule(ax, y)
        y -= 0.025

        # ── Analysis Summary ──────────────────────────────────────────
        y = _section_header(ax, y, 'Analysis Summary')

        bullets = []
        explanations = []
        for k, v in summary.items():
            if isinstance(v, list):
                v = ', '.join(map(str, v[:10])) + (' ...' if len(v) > 10 else '')
            bullets.append((k, str(v)))
            desc = explain_check(k)
            if desc:
                explanations.append((k, desc))

        for label, value in bullets:
            combined = f'{label}:  {value}'
            for wrapped in textwrap.wrap(combined, width=95):
                fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H)
                ax.text(MARGIN_L + 0.01, y, wrapped, ha='left', va='top',
                        fontsize=9.5, color=C_TEXT)
                y -= LINE_H

        y -= 0.02
        _draw_rule(ax, y)
        y -= 0.025

        # ── Check Definitions ─────────────────────────────────────────
        fig, ax, y = _need_space(pdf, fig, ax, y, 0.06, 'Check Definitions')
        y = _section_header(ax, y, 'Check Definitions')

        for check_name, check_desc in explanations:
            fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H * 2)
            ax.text(MARGIN_L + 0.01, y, check_name, ha='left', va='top',
                    fontsize=9, weight='bold', color=C_ACCENT)
            y -= LINE_H
            for wrapped in textwrap.wrap(check_desc, width=95):
                fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H_SM)
                ax.text(MARGIN_L + 0.03, y, wrapped, ha='left', va='top',
                        fontsize=8.5, color=C_TEXT)
                y -= LINE_H_SM
            y -= 0.005

        y -= 0.02
        _draw_rule(ax, y)
        y -= 0.025

        # ── Vintage-Level Issues ──────────────────────────────────────
        fig, ax, y = _need_space(pdf, fig, ax, y, 0.14, 'Vintage-Level Issues')
        y = _section_header(ax, y, 'Vintage-Level Data Quality Findings')

        intro = (
            "A vintage represents a closed cohort of loans grouped by origination quarter "
            "(e.g., 2023-Q1 comprises all loans originated January through March 2023). "
            "The validations below assess cohort-level data consistency, ensuring denominator "
            "stability, adequate initial coverage, and monotonicity of cumulative default rates."
        )
        for wrapped in textwrap.wrap(intro, width=100):
            fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H)
            ax.text(MARGIN_L + 0.01, y, wrapped, ha='left', va='top',
                    fontsize=9.5, style='italic', color=C_MUTED)
            y -= LINE_H
        y -= 0.02

        has_vintage_issues = vintage_issues_df is not None and not vintage_issues_df.empty

        if has_vintage_issues:
            issue_types = vintage_issues_df['issue'].unique().tolist()
            issue_num = 0

            for issue_type in issue_types:
                issue_num += 1
                fig, ax, y = _need_space(pdf, fig, ax, y, 0.12, issue_type)

                # Issue sub-header with numbering and colored left bar
                ax.plot([MARGIN_L, MARGIN_L], [y + 0.005, y - 0.025],
                        color=C_RED, lw=3, solid_capstyle='round')
                ax.text(MARGIN_L + 0.015, y, f'{issue_num}.  {issue_type}',
                        ha='left', va='top', fontsize=11, weight='bold', color=C_RED)
                y -= 0.035

                # Corporate explanation
                corp_text = VINTAGE_ISSUE_EXPLANATIONS.get(issue_type, '')
                if corp_text:
                    for wrapped in textwrap.wrap(corp_text, width=95):
                        fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H)
                        ax.text(MARGIN_L + 0.03, y, wrapped, ha='left', va='top',
                                fontsize=9, color=C_TEXT)
                        y -= LINE_H
                    y -= 0.01

                # Affected vintages
                sub = vintage_issues_df[vintage_issues_df['issue'] == issue_type]
                affected = sub['Vintage'].unique().tolist() if 'Vintage' in sub.columns else []
                if affected:
                    vlist = ', '.join(str(v) for v in affected[:20])
                    if len(affected) > 20:
                        vlist += ' ...'
                    label_line = f'Affected vintages ({len(affected)}):  {vlist}'
                    for wrapped in textwrap.wrap(label_line, width=95):
                        fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H)
                        ax.text(MARGIN_L + 0.03, y, wrapped, ha='left', va='top',
                                fontsize=9, weight='bold', color=C_DARK)
                        y -= LINE_H
                    y -= 0.005

                # Sample data table
                detail_cols = [c for c in sub.columns if c != 'issue']
                if detail_cols:
                    sample = sub[detail_cols].head(10)

                    # Table header background
                    fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H + 0.005)
                    ax.add_patch(FancyBboxPatch((MARGIN_L + 0.025, y - 0.013),
                                 MARGIN_R - MARGIN_L - 0.035, 0.018,
                                 boxstyle="square,pad=0", facecolor=C_BG_LIGHT,
                                 edgecolor=C_RULE, linewidth=0.5))
                    header = '    '.join(str(c) for c in detail_cols)
                    ax.text(MARGIN_L + 0.035, y, header, ha='left', va='top',
                            fontsize=7.5, weight='bold', color=C_DARK,
                            fontfamily='monospace')
                    y -= LINE_H_SM

                    for _, row in sample.iterrows():
                        vals = '    '.join(str(row[c]) for c in detail_cols)
                        for wrapped in textwrap.wrap(vals, width=110):
                            fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H_SM)
                            ax.text(MARGIN_L + 0.035, y, wrapped, ha='left', va='top',
                                    fontsize=7.5, color=C_TEXT,
                                    fontfamily='monospace')
                            y -= LINE_H_SM

                y -= 0.02
                _draw_rule(ax, y, color=C_RULE, lw=0.3)
                y -= 0.015
        else:
            fig, ax, y = _need_space(pdf, fig, ax, y, LINE_H)
            ax.text(MARGIN_L + 0.01, y, 'No vintage-level issues detected.',
                    ha='left', va='top', fontsize=11, color=C_GREEN, weight='bold')
            y -= LINE_H

        # Footer
        ax.text(0.5, 0.02, 'Confidential  |  For Internal Use Only',
                ha='center', va='bottom', fontsize=7.5, color=C_MUTED,
                style='italic')

        _save_page(pdf, fig)
    return buf.getvalue()


def export_issues_excel(issues_df: pd.DataFrame, vintage_issues: pd.DataFrame) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine='xlsxwriter') as xw:
        (issues_df if (issues_df is not None and not issues_df.empty)
         else pd.DataFrame({'note':['No row-level issues sampled']})).to_excel(xw, index=False, sheet_name='Row issues')
        (vintage_issues if (vintage_issues is not None and not vintage_issues.empty)
         else pd.DataFrame({'note':['No vintage-level issues']})).to_excel(xw, index=False, sheet_name='Vintage issues')
    return out.getvalue()

def export_disappeared_loans_excel(disappeared_df: pd.DataFrame) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine='xlsxwriter') as xw:
        (disappeared_df if (disappeared_df is not None and not disappeared_df.empty)
         else pd.DataFrame({'note': ['No disappeared loans detected']})).to_excel(xw, index=False, sheet_name='Disappeared Loans')
    return out.getvalue()

def export_consistency_excel(summary_df: pd.DataFrame,
                             chart_wide: pd.DataFrame,
                             gran_key: str) -> bytes:
    """Excel workbook with three sheets for table-vs-chart consistency check.

    Sheet 1 – Table Summary: the vintage default summary (Cum_PD, etc.)
    Sheet 2 – Chart Data: full chart series for every vintage
    Sheet 3 – Comparison: side-by-side Cum_PD from the table vs the endpoint
              of each vintage series in the chart
    """
    out = BytesIO()
    period_col = chart_wide.index.name or gran_key

    with pd.ExcelWriter(out, engine='xlsxwriter') as xw:
        # Sheet 1 — Table summary
        tbl = summary_df.copy()
        tbl['Cum PD (%)'] = tbl['Cum_PD'] * 100
        tbl['Annualized default rate (%)'] = tbl['Default_rate_pa'] * 100
        tbl.to_excel(xw, index=False, sheet_name='Table Summary')

        # Sheet 2 — Chart data (periods in rows, vintages in columns)
        chart_exp = chart_wide.reset_index()
        if gran_key == 'QOB':
            chart_exp.insert(0, 'Months', chart_exp[period_col] * 3)
        chart_exp.to_excel(xw, index=False, sheet_name='Chart Data')

        # Sheet 3 — Comparison: table Cum_PD vs chart endpoint per vintage
        rows = []
        for _, r in summary_df.iterrows():
            v = r['Vintage']
            table_cum_pd = r['Cum_PD']
            if v in chart_wide.columns:
                series = chart_wide[v].dropna()
                chart_endpoint = series.iloc[-1] if len(series) > 0 else None
                last_period = int(series.index[-1]) if len(series) > 0 else None
            else:
                chart_endpoint = None
                last_period = None
            diff = (chart_endpoint - table_cum_pd) if chart_endpoint is not None else None
            rows.append({
                'Vintage': v,
                'Unique loans': r['Unique_loans'],
                'Defaulted loans': r['Defaulted_loans'],
                'Table Cum PD': table_cum_pd,
                'Table Cum PD (%)': table_cum_pd * 100,
                f'Chart Endpoint ({gran_key})': last_period,
                'Chart Endpoint PD': chart_endpoint,
                'Chart Endpoint PD (%)': chart_endpoint * 100 if chart_endpoint is not None else None,
                'Difference (abs)': diff,
                'Match': 'Yes' if diff is not None and abs(diff) < 1e-6 else 'No',
            })
        comp_df = pd.DataFrame(rows)
        comp_df.to_excel(xw, index=False, sheet_name='Comparison')

        # Format the Comparison sheet
        wb = xw.book
        ws = xw.sheets['Comparison']
        pct_fmt = wb.add_format({'num_format': '0.00%'})
        pct_cols = ['Table Cum PD', 'Chart Endpoint PD', 'Difference (abs)']
        for col_name in pct_cols:
            if col_name in comp_df.columns:
                col_idx = comp_df.columns.get_loc(col_name)
                ws.set_column(col_idx, col_idx, 14, pct_fmt)

    return out.getvalue()

# ──────────────────────────────────────────────────────────────────────────────
# Vintage default summary (Cum_PD, Obs_Time, annualized rate)
# ──────────────────────────────────────────────────────────────────────────────
def compute_vintage_default_summary(raw_df: pd.DataFrame, dpd_threshold: int,
                                    pd_by_amount: bool = False,
                                    vintage_granularity: str = 'Q') -> pd.DataFrame:
    dfn = normalize_columns(raw_df); dfn = ensure_types(dfn)
    dfn = add_vintage_mob(dfn, granularity=vintage_granularity)
    dfn = dfn.sort_values(['Loan ID','Observation date'])
    dfn['__def'] = (dfn['Days past due'] >= dpd_threshold)

    g = dfn.groupby('Loan ID', sort=False)
    first_obs = g['Observation date'].min()
    last_obs  = g['Observation date'].max()
    first_vintage = g['Vintage'].first()
    first_def_date = (dfn.loc[dfn['__def']]
                        .groupby('Loan ID', sort=False)['Observation date']
                        .min())

    loan_df = pd.DataFrame({'Vintage': first_vintage, 'first_obs': first_obs, 'last_obs': last_obs})
    loan_df['def_date']  = first_def_date
    loan_df['defaulted'] = loan_df['def_date'].notna()
    loan_df['obs_end']   = loan_df['def_date'].fillna(loan_df['last_obs'])

    obs_days = (loan_df['obs_end'] - loan_df['first_obs']).dt.days
    loan_df['Obs_Time_years'] = (obs_days / 365.25).astype('float32')
    loan_df.loc[loan_df['Obs_Time_years'] <= 0, 'Obs_Time_years'] = np.nan

    if pd_by_amount:
        # Get origination amount per loan (first observation)
        loan_orig_amt = g['Origination amount'].first()
        loan_df['Origination_amount'] = loan_orig_amt

        # Get current amount at first default date for each defaulted loan
        def_rows = dfn.loc[dfn['__def']].copy()
        first_def_rows = (def_rows.sort_values('Observation date')
                          .drop_duplicates(subset='Loan ID', keep='first'))
        first_def_cur_amt = first_def_rows.set_index('Loan ID')['Current amount']
        loan_df['Default_current_amount'] = first_def_cur_amt
        loan_df.loc[~loan_df['defaulted'], 'Default_current_amount'] = 0.0

        out = (loan_df.groupby('Vintage', as_index=False)
               .agg(Unique_loans=('Vintage', 'size'),
                    Defaulted_loans=('defaulted', 'sum'),
                    Total_Origination_Amount=('Origination_amount', 'sum'),
                    Total_Default_Amount=('Default_current_amount', 'sum'),
                    Observation_Time=('Obs_Time_years', 'median')))
        out['Cum_PD'] = np.where(
            out['Total_Origination_Amount'] > 0,
            out['Total_Default_Amount'] / out['Total_Origination_Amount'],
            0.0
        ).astype('float32')
    else:
        out = (loan_df.groupby('Vintage', as_index=False)
               .agg(Unique_loans=('Vintage','size'),
                    Defaulted_loans=('defaulted','sum'),
                    Cum_PD=('defaulted','mean'),
                    Observation_Time=('Obs_Time_years','median')))

    m = out['Observation_Time'] > 0
    out['Default_rate_pa'] = np.nan
    out.loc[m, 'Default_rate_pa'] = 1 - np.power(1 - out.loc[m, 'Cum_PD'], 1 / out.loc[m, 'Observation_Time'])
    return out.sort_values('Vintage').reset_index(drop=True)

# ──────────────────────────────────────────────────────────────────────────────
# Segment-based PD summary (group by any non-reserved column)
# ──────────────────────────────────────────────────────────────────────────────
def compute_segment_default_summary(raw_df: pd.DataFrame, dpd_threshold: int,
                                    segment_col: str,
                                    pd_by_amount: bool = False,
                                    vintage_granularity: str = 'Q') -> pd.DataFrame:
    """Compute PD metrics grouped by values of *segment_col* instead of Vintage."""
    dfn = normalize_columns(raw_df); dfn = ensure_types(dfn)
    dfn = add_vintage_mob(dfn, granularity=vintage_granularity)

    # Preserve original segment values per loan (use the value from the first observation)
    if segment_col not in dfn.columns:
        raise KeyError(f"Column '{segment_col}' not found in the data after normalisation.")

    dfn = dfn.sort_values(['Loan ID', 'Observation date'])
    dfn['__def'] = (dfn['Days past due'] >= dpd_threshold)

    g = dfn.groupby('Loan ID', sort=False)
    first_obs      = g['Observation date'].min()
    last_obs       = g['Observation date'].max()
    first_segment  = g[segment_col].first()
    first_def_date = (dfn.loc[dfn['__def']]
                        .groupby('Loan ID', sort=False)['Observation date']
                        .min())

    loan_df = pd.DataFrame({
        'Segment': first_segment,
        'first_obs': first_obs,
        'last_obs': last_obs,
    })
    loan_df['def_date']  = first_def_date
    loan_df['defaulted'] = loan_df['def_date'].notna()
    loan_df['obs_end']   = loan_df['def_date'].fillna(loan_df['last_obs'])

    obs_days = (loan_df['obs_end'] - loan_df['first_obs']).dt.days
    loan_df['Obs_Time_years'] = (obs_days / 365.25).astype('float32')
    loan_df.loc[loan_df['Obs_Time_years'] <= 0, 'Obs_Time_years'] = np.nan

    if pd_by_amount:
        loan_orig_amt = g['Origination amount'].first()
        loan_df['Origination_amount'] = loan_orig_amt

        def_rows = dfn.loc[dfn['__def']].copy()
        first_def_rows = (def_rows.sort_values('Observation date')
                          .drop_duplicates(subset='Loan ID', keep='first'))
        first_def_cur_amt = first_def_rows.set_index('Loan ID')['Current amount']
        loan_df['Default_current_amount'] = first_def_cur_amt
        loan_df.loc[~loan_df['defaulted'], 'Default_current_amount'] = 0.0

        out = (loan_df.groupby('Segment', as_index=False)
               .agg(Unique_loans=('Segment', 'size'),
                    Defaulted_loans=('defaulted', 'sum'),
                    Total_Origination_Amount=('Origination_amount', 'sum'),
                    Total_Default_Amount=('Default_current_amount', 'sum'),
                    Observation_Time=('Obs_Time_years', 'median')))
        out['Cum_PD'] = np.where(
            out['Total_Origination_Amount'] > 0,
            out['Total_Default_Amount'] / out['Total_Origination_Amount'],
            0.0
        ).astype('float32')
    else:
        out = (loan_df.groupby('Segment', as_index=False)
               .agg(Unique_loans=('Segment', 'size'),
                    Defaulted_loans=('defaulted', 'sum'),
                    Cum_PD=('defaulted', 'mean'),
                    Observation_Time=('Obs_Time_years', 'median')))

    m = out['Observation_Time'] > 0
    out['Default_rate_pa'] = np.nan
    out.loc[m, 'Default_rate_pa'] = 1 - np.power(1 - out.loc[m, 'Cum_PD'], 1 / out.loc[m, 'Observation_Time'])
    return out.sort_values('Segment').reset_index(drop=True)

# ──────────────────────────────────────────────────────────────────────────────
# Loan-level builder & target-PD filter recommender
# ──────────────────────────────────────────────────────────────────────────────
# Internal loan-level columns that should NOT be offered as eligibility filters
_LOAN_INTERNAL_COLS = {
    'Loan ID', 'first_obs', 'last_obs', 'def_date', 'obs_end',
    'defaulted', 'Obs_Time_years', 'Origination_amount',
    'Default_current_amount',
}


def compute_loan_level_table(raw_df: pd.DataFrame, dpd_threshold: int,
                             vintage_granularity: str = 'Q') -> pd.DataFrame:
    """Build a one-row-per-loan table with default flag, observation time, origination
    amount, default current amount, and every non-reserved attribute (value at the
    first observation). Used as the working set for the PD Optimizer."""
    dfn = normalize_columns(raw_df); dfn = ensure_types(dfn)
    dfn = add_vintage_mob(dfn, granularity=vintage_granularity)
    dfn = dfn.sort_values(['Loan ID', 'Observation date'])
    dfn['__def'] = (dfn['Days past due'] >= dpd_threshold)

    g = dfn.groupby('Loan ID', sort=False)
    first_obs = g['Observation date'].min()
    last_obs  = g['Observation date'].max()
    orig_amt  = g['Origination amount'].first()
    first_def_date = (dfn.loc[dfn['__def']]
                        .groupby('Loan ID', sort=False)['Observation date']
                        .min())

    loan_df = pd.DataFrame({
        'first_obs': first_obs,
        'last_obs':  last_obs,
        'Origination_amount': orig_amt,
    })
    loan_df['def_date']  = first_def_date
    loan_df['defaulted'] = loan_df['def_date'].notna()
    loan_df['obs_end']   = loan_df['def_date'].fillna(loan_df['last_obs'])
    obs_days = (loan_df['obs_end'] - loan_df['first_obs']).dt.days
    loan_df['Obs_Time_years'] = (obs_days / 365.25).astype('float64')
    loan_df.loc[loan_df['Obs_Time_years'] <= 0, 'Obs_Time_years'] = np.nan

    # Current amount at first default date (0 for non-defaulted loans)
    def_rows = dfn.loc[dfn['__def']]
    if not def_rows.empty:
        first_def_rows = (def_rows.sort_values('Observation date')
                          .drop_duplicates(subset='Loan ID', keep='first'))
        first_def_cur_amt = first_def_rows.set_index('Loan ID')['Current amount']
    else:
        first_def_cur_amt = pd.Series(dtype='float64')
    loan_df['Default_current_amount'] = first_def_cur_amt
    loan_df.loc[~loan_df['defaulted'], 'Default_current_amount'] = 0.0
    loan_df['Default_current_amount'] = loan_df['Default_current_amount'].fillna(0.0)

    # Attach every non-reserved attribute (first-observation value per loan)
    non_reserved = [c for c in dfn.columns
                    if str(c).strip() not in RESERVED_COLS
                    and not str(c).startswith('__')]
    for c in non_reserved:
        loan_df[c] = g[c].first()

    return loan_df.reset_index()


def _ann_pd_from_subset(sub: pd.DataFrame, pd_by_amount: bool) -> tuple:
    """Return (cum_pd, ann_pd, n_loans, orig_amt) for a loan-level subset."""
    n = len(sub)
    if n == 0:
        return (np.nan, np.nan, 0, 0.0)
    defaulted = int(sub['defaulted'].sum())
    orig_amt  = float(sub['Origination_amount'].sum(skipna=True))
    if pd_by_amount:
        def_amt = float(sub['Default_current_amount'].sum(skipna=True))
        cum_pd  = (def_amt / orig_amt) if orig_amt > 0 else np.nan
    else:
        cum_pd  = defaulted / n
    obs = sub['Obs_Time_years'].dropna()
    obs_time = float(obs.median()) if not obs.empty else np.nan
    if (cum_pd is None) or np.isnan(cum_pd) or np.isnan(obs_time) \
            or obs_time <= 0 or cum_pd >= 1:
        ann_pd = cum_pd
    else:
        ann_pd = 1 - (1 - cum_pd) ** (1 / obs_time)
    return (cum_pd, ann_pd, n, orig_amt)


def suggest_filters_for_target_pd(
    loan_df: pd.DataFrame,
    target_ann_pd: float,
    candidate_cols: list,
    pd_by_amount: bool = False,
    n_quantiles: int = 21,
    min_retention: float = 0.05,
    required_map: Optional[dict] = None,
) -> tuple:
    """For each candidate column, find the single eligibility filter that brings the
    portfolio's annualized PD at or below *target_ann_pd* while retaining the most
    loans (or origination amount when *pd_by_amount*).

    Numeric columns are searched as one-sided thresholds (``col ≤ v`` or ``col ≥ v``)
    across *n_quantiles* quantile cut-points. Categorical columns are searched by
    greedily excluding the highest-PD categories.

    *required_map* — optional ``{col: {'type': 'numeric', 'lo': x, 'hi': y}}`` or
    ``{col: {'type': 'categorical', 'values': [...]}}``. Suggestions on those
    columns must preserve the listed range / values (the reported filter is
    relaxed to the intersection so the constraint is not violated).

    Returns ``(feasible_df, partial_df, baseline)`` where *feasible_df* holds filters
    that reach target, *partial_df* holds the best progress per column when no filter
    reaches target, and *baseline* is ``(cum_pd, ann_pd, n_loans, orig_amt)`` of the
    input portfolio.
    """
    required_map = required_map or {}
    baseline = _ann_pd_from_subset(loan_df, pd_by_amount)
    _, base_ann_pd, base_n, base_amt = baseline

    feasible_rows = []
    partial_rows  = []

    if base_n == 0 or np.isnan(base_ann_pd):
        return pd.DataFrame(), pd.DataFrame(), baseline

    qs = np.linspace(0.0, 1.0, n_quantiles)

    def _record(col: str, ftype: str, filter_desc: str,
                cp, ap, rl, ra, extras: dict | None = None) -> dict:
        row = {
            'Column': col,
            'Type': ftype,
            'Filter': filter_desc,
            'Cum PD (%)': (cp * 100) if cp is not None and not (isinstance(cp, float) and np.isnan(cp)) else np.nan,
            'Annualized PD (%)': ap * 100 if ap is not None and not np.isnan(ap) else np.nan,
            'Retained loans': rl,
            'Retained origination amt': ra,
            'Retention (count) %': (rl / base_n * 100) if base_n > 0 else 0.0,
            'Retention (amt) %':   (ra / base_amt * 100) if base_amt > 0 else 0.0,
        }
        if extras:
            row.update(extras)
        return row

    for col in candidate_cols:
        if col not in loan_df.columns:
            continue
        col_series = loan_df[col]
        is_num = pd.api.types.is_numeric_dtype(col_series) and col_series.nunique(dropna=True) > 1

        best_feasible = None     # (retained, row_dict)
        best_progress = None     # (ann_pd, row_dict) — track smallest ann_pd seen

        def _consider(ap, rl, ra, row_dict):
            """Decide whether this candidate beats the running bests."""
            nonlocal best_feasible, best_progress
            if ap is None or np.isnan(ap):
                return
            retained = ra if pd_by_amount else rl
            min_keep = (base_amt if pd_by_amount else base_n) * min_retention
            if retained < min_keep:
                return
            if ap <= target_ann_pd:
                if best_feasible is None or retained > best_feasible[0]:
                    best_feasible = (retained, row_dict)
            if best_progress is None or ap < best_progress[0]:
                best_progress = (ap, row_dict)

        col_constraint = required_map.get(col)

        if is_num:
            numeric_vals = pd.to_numeric(col_series, errors='coerce')
            if numeric_vals.dropna().empty:
                continue
            try:
                val_grid = numeric_vals.quantile(qs).values
            except Exception:
                continue

            # Constraint: if a required range is given, threshold must preserve it.
            req_lo = req_hi = None
            if col_constraint and col_constraint.get('type') == 'numeric':
                req_lo = col_constraint.get('lo')
                req_hi = col_constraint.get('hi')

            # Upper threshold: keep col ≤ v  → v must be ≥ req_hi
            for q, v in zip(qs, val_grid):
                if q == 0 or np.isnan(v):
                    continue
                if req_hi is not None and v < req_hi:
                    continue
                sub = loan_df[numeric_vals <= v]
                cp, ap, rl, ra = _ann_pd_from_subset(sub, pd_by_amount)
                row = _record(col, 'Numeric ≤',
                              f'{col} ≤ {v:.6g}',
                              cp, ap, rl, ra,
                              extras={'Threshold direction': '≤',
                                      'Threshold value':     float(v)})
                _consider(ap, rl, ra, row)

            # Lower threshold: keep col ≥ v  → v must be ≤ req_lo
            for q, v in zip(qs, val_grid):
                if q == 1 or np.isnan(v):
                    continue
                if req_lo is not None and v > req_lo:
                    continue
                sub = loan_df[numeric_vals >= v]
                cp, ap, rl, ra = _ann_pd_from_subset(sub, pd_by_amount)
                row = _record(col, 'Numeric ≥',
                              f'{col} ≥ {v:.6g}',
                              cp, ap, rl, ra,
                              extras={'Threshold direction': '≥',
                                      'Threshold value':     float(v)})
                _consider(ap, rl, ra, row)
        else:
            # Categorical — fill missing as "(missing)" for stable grouping
            tmp_col = col_series.astype('object').where(col_series.notna(), other='(missing)')
            aux = pd.DataFrame({
                'cat': tmp_col.values,
                'defaulted': loan_df['defaulted'].values,
                'orig':      loan_df['Origination_amount'].values,
                'def_amt':   loan_df['Default_current_amount'].values,
            })
            agg = aux.groupby('cat', dropna=False).agg(
                n=('defaulted', 'size'),
                d=('defaulted', 'sum'),
                orig=('orig', 'sum'),
                def_amt=('def_amt', 'sum'),
            )
            if pd_by_amount:
                agg['cat_pd'] = np.where(agg['orig'] > 0, agg['def_amt'] / agg['orig'], 0.0)
            else:
                agg['cat_pd'] = np.where(agg['n'] > 0, agg['d'] / agg['n'], 0.0)
            agg = agg.sort_values('cat_pd', ascending=False)

            # Constraint: if required values are listed, never drop them.
            required_vals: set = set()
            if col_constraint and col_constraint.get('type') == 'categorical':
                required_vals = set(col_constraint.get('values') or [])

            ordered = list(agg.index)
            if len(ordered) < 2:
                continue  # nothing to drop

            # Iteratively drop highest-PD category (skipping required categories)
            for drop_n in range(len(ordered)):
                kept_cats = [c2 for c2 in ordered[drop_n:]]
                # Always force required categories to be present
                for rv in required_vals:
                    if rv in ordered and rv not in kept_cats:
                        kept_cats.append(rv)
                if not kept_cats:
                    break
                mask = tmp_col.isin(kept_cats)
                sub = loan_df[mask.values]
                cp, ap, rl, ra = _ann_pd_from_subset(sub, pd_by_amount)
                preview = ', '.join(str(v) for v in kept_cats[:6])
                if len(kept_cats) > 6:
                    preview += f', … (+{len(kept_cats) - 6} more)'
                dropped_cats = [c2 for c2 in ordered if c2 not in kept_cats]
                row = _record(col, 'Categorical',
                              f'{col} ∈ {{{preview}}}',
                              cp, ap, rl, ra,
                              extras={'Kept categories': len(kept_cats),
                                      'Dropped categories': len(ordered) - len(kept_cats),
                                      'Kept values': kept_cats,
                                      'Dropped values': dropped_cats,
                                      'Total categories': len(ordered)})
                _consider(ap, rl, ra, row)
                # Once only required categories remain, further dropping would
                # violate the constraint — stop.
                if required_vals and set(ordered[drop_n + 1:]).issubset(required_vals):
                    break

        if best_feasible is not None:
            feasible_rows.append(best_feasible[1])
        elif best_progress is not None:
            partial_rows.append(best_progress[1])

    sort_key = 'Retained origination amt' if pd_by_amount else 'Retained loans'

    feasible_df = pd.DataFrame(feasible_rows)
    if not feasible_df.empty:
        feasible_df = feasible_df.sort_values(sort_key, ascending=False).reset_index(drop=True)

    partial_df = pd.DataFrame(partial_rows)
    if not partial_df.empty:
        partial_df = partial_df.sort_values('Annualized PD (%)', ascending=True).reset_index(drop=True)

    return feasible_df, partial_df, baseline


def suggest_combo_filter(
    loan_df: pd.DataFrame,
    target_ann_pd: float,
    feasible_df: pd.DataFrame,
    partial_df: pd.DataFrame,
    pd_by_amount: bool,
    top_k: int = 3,
) -> Optional[dict]:
    """Attempt to combine the top single-column filters (from *feasible_df* first,
    else from *partial_df*) into a two-filter rule that meets target with higher
    retention. Returns a dict describing the combo, or None if no improvement."""
    if loan_df.empty:
        return None

    pool = feasible_df if not feasible_df.empty else partial_df
    if pool.empty or len(pool) < 2:
        return None

    pool = pool.head(top_k)
    baseline = _ann_pd_from_subset(loan_df, pd_by_amount)
    _, base_ann_pd, base_n, base_amt = baseline

    def _apply_filter(df, row):
        col = row['Column']
        if col not in df.columns:
            return None
        if row['Type'].startswith('Numeric'):
            vals = pd.to_numeric(df[col], errors='coerce')
            v = row.get('Threshold value')
            d = row.get('Threshold direction')
            if v is None or d is None:
                return None
            return df[(vals <= v) if d == '≤' else (vals >= v)]
        else:
            kept = row.get('Kept values')
            if not isinstance(kept, list):
                return None
            ser = df[col].astype('object').where(df[col].notna(), other='(missing)')
            return df[ser.isin(kept)]

    best = None  # (retained, combo_dict)
    for i in range(len(pool)):
        for j in range(i + 1, len(pool)):
            r1 = pool.iloc[i].to_dict()
            r2 = pool.iloc[j].to_dict()
            if r1['Column'] == r2['Column']:
                continue
            step1 = _apply_filter(loan_df, r1)
            if step1 is None or step1.empty:
                continue
            step2 = _apply_filter(step1, r2)
            if step2 is None or step2.empty:
                continue
            cp, ap, rl, ra = _ann_pd_from_subset(step2, pd_by_amount)
            if ap is None or np.isnan(ap):
                continue
            retained = ra if pd_by_amount else rl
            if ap <= target_ann_pd and (best is None or retained > best[0]):
                best = (retained, {
                    'Filter 1': r1['Filter'],
                    'Filter 2': r2['Filter'],
                    'Cum PD (%)': (cp * 100) if cp is not None and not np.isnan(cp) else np.nan,
                    'Annualized PD (%)': ap * 100,
                    'Retained loans': rl,
                    'Retained origination amt': ra,
                    'Retention (count) %': (rl / base_n * 100) if base_n > 0 else 0.0,
                    'Retention (amt) %':   (ra / base_amt * 100) if base_amt > 0 else 0.0,
                })
    return best[1] if best else None


# ──────────────────────────────────────────────────────────────────────────────
# UI - SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    # Sidebar branding - light background with dark text
    st.markdown(
        f"""
        <div style="
            text-align: center;
            padding: 16px 0 24px 0;
            border-bottom: 1px solid {WB_BORDER};
            margin-bottom: 24px;
        ">
            <div style="
                background: {WB_PRIMARY};
                width: 56px;
                height: 56px;
                border-radius: 14px;
                display: flex;
                align-items: center;
                justify-content: center;
                margin: 0 auto 12px auto;
            ">
                <svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2">
                    <path d="M12 2L2 7l10 5 10-5-10-5z"/>
                    <path d="M2 17l10 5 10-5"/>
                    <path d="M2 12l10 5 10-5"/>
                </svg>
            </div>
            <h3 style="margin: 0; font-size: 1.1rem; font-weight: 600; color: {WB_PRIMARY};">Analytics Suite</h3>
            <p style="margin: 4px 0 0 0; font-size: 0.75rem; color: {WB_MUTED};">Vintage Curve Analysis</p>
        </div>
        """,
        unsafe_allow_html=True
    )

    # Quick start guide
    with st.expander("📘 Quick Start Guide", expanded=False):
        st.markdown(
            """
            **Step 1:** Configure your default threshold

            **Step 2:** Upload your Excel file (.xlsx)

            **Step 3:** Select sheet and click Load

            **Step 4:** Explore the analysis tabs
            """
        )

    st.markdown("---")

    # Live memory indicator
    try:
        import resource
        _mem_mb = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss / 1024
        _mem_color = WB_SUCCESS if _mem_mb < 700 else ("#F59E0B" if _mem_mb < 900 else "#EF4444")
        st.caption(f'Memory: **{_mem_mb:,.0f} MB** <span style="color:{_mem_color};">●</span>',
                   unsafe_allow_html=True)
    except Exception:
        pass

# ──────────────────────────────────────────────────────────────────────────────
# UI - MAIN CONTENT
# ──────────────────────────────────────────────────────────────────────────────
left, right = st.columns([1, 3], gap="large")

with left:
    # Settings Card
    st.markdown(
        f"""
        <div style="
            background: white;
            border-radius: 16px;
            padding: 24px;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
            border: 1px solid {WB_BORDER};
            margin-bottom: 24px;
        ">
            <div style="display: flex; align-items: center; gap: 12px; margin-bottom: 20px;">
                <div style="
                    background: linear-gradient(135deg, {WB_PRIMARY} 0%, {WB_SECONDARY} 100%);
                    width: 40px;
                    height: 40px;
                    border-radius: 10px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                ">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2">
                        <circle cx="12" cy="12" r="3"/>
                        <path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1 0 2.83 2 2 0 0 1-2.83 0l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-2 2 2 2 0 0 1-2-2v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83 0 2 2 0 0 1 0-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1-2-2 2 2 0 0 1 2-2h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 0-2.83 2 2 0 0 1 2.83 0l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 2-2 2 2 0 0 1 2 2v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 0 2 2 0 0 1 0 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 2 2 2 2 0 0 1-2 2h-.09a1.65 1.65 0 0 0-1.51 1z"/>
                    </svg>
                </div>
                <div>
                    <h3 style="margin: 0; font-size: 1.1rem; color: {WB_TEXT};">Configuration</h3>
                    <p style="margin: 0; font-size: 0.8rem; color: {WB_MUTED};">Analysis parameters</p>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    dpd_threshold = st.number_input(
        '🎯 Default Threshold (DPD ≥)',
        min_value=1,
        max_value=365,
        value=90,
        step=1,
        help='Loans are considered in default when Days Past Due exceeds this threshold'
    )

    vintage_granularity_label = st.selectbox(
        '📅 Vintage Grouping',
        list(VINTAGE_GRANULARITY_OPTIONS.keys()),
        index=1,  # Quarterly default
        help='Group origination cohorts by month, quarter, 6 months, or year.'
    )
    vintage_granularity = VINTAGE_GRANULARITY_OPTIONS[vintage_granularity_label]

    pretty_ints = st.checkbox(
        '📊 Format with thousand separators',
        value=False,
        help='Display numbers as 12,345 instead of 12345'
    )

    st.markdown("---")

    # Upload Card
    st.markdown(
        f"""
        <div style="
            background: white;
            border-radius: 16px;
            padding: 24px;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
            border: 1px solid {WB_BORDER};
            margin-bottom: 16px;
        ">
            <div style="display: flex; align-items: center; gap: 12px; margin-bottom: 16px;">
                <div style="
                    background: linear-gradient(135deg, {WB_SUCCESS} 0%, #059669 100%);
                    width: 40px;
                    height: 40px;
                    border-radius: 10px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                ">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2">
                        <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>
                        <polyline points="17 8 12 3 7 8"/>
                        <line x1="12" y1="3" x2="12" y2="15"/>
                    </svg>
                </div>
                <div>
                    <h3 style="margin: 0; font-size: 1.1rem; color: {WB_TEXT};">Data Import</h3>
                    <p style="margin: 0; font-size: 0.8rem; color: {WB_MUTED};">Upload your Excel file</p>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    uploaded = st.file_uploader(
        'Select Excel file',
        type=['xlsx'],
        accept_multiple_files=False,
        help='Upload a .xlsx file containing loan portfolio data'
    )

    # Persist full dataset and flagged row indices
    if 'df_full' not in st.session_state:
        st.session_state['df_full'] = None
    if 'flagged_indices' not in st.session_state:
        st.session_state['flagged_indices'] = set()

    if uploaded:
        size_mb = uploaded.size / (1024 * 1024)

        # File info display
        st.markdown(
            f"""
            <div style="
                background: {WB_LIGHT};
                border-radius: 10px;
                padding: 12px 16px;
                margin: 12px 0;
                border-left: 4px solid {WB_PRIMARY};
            ">
                <div style="font-size: 0.9rem; font-weight: 600; color: {WB_TEXT};">📁 {uploaded.name}</div>
                <div style="font-size: 0.8rem; color: {WB_MUTED};">Size: {size_mb:,.2f} MB</div>
            </div>
            """,
            unsafe_allow_html=True
        )

        if size_mb > MAX_MB:
            st.warning('⚠️ Large file detected. Consider filtering columns or using CSV format.')

        from openpyxl import load_workbook
        names = load_workbook(filename=BytesIO(uploaded.getvalue()), read_only=True, data_only=True).sheetnames
        sheet = st.selectbox('📋 Select worksheet', options=names, index=0)
        header_row = st.number_input('📍 Header row position', min_value=1, value=1, step=1, help='Row 1 = first row')

        if st.button('🚀 Load Dataset', type='primary', use_container_width=True):
            with st.status('Loading dataset...', expanded=True) as status:
                try:
                    st.write("📖 Reading Excel file...")
                    logger.info("Loading file: %s (%.2f MB)", uploaded.name, size_mb)
                    df_full = load_full(uploaded.getvalue(), sheet=sheet, header=header_row - 1)
                    logger.info("Loaded %d rows x %d cols", len(df_full), len(df_full.columns))
                    _log_resource_usage("after load_full")
                    st.write(f"✅ Loaded {len(df_full):,} rows and {len(df_full.columns)} columns")
                    st.session_state['df_full'] = df_full
                    gc.collect()
                    status.update(label='✅ Dataset loaded successfully!', state='complete')
                except Exception:
                    logger.exception("CRASH during dataset load")
                    status.update(label='❌ Load failed', state='error')
                    st.error(f"Failed to load file:\n```\n{traceback.format_exc()}\n```")
with right:
    if st.session_state['df_full'] is not None:
        chosen_df_raw = st.session_state['df_full']

        # Dataset summary metrics
        st.markdown(
            f"""
            <div style="
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 16px;
                margin-bottom: 24px;
            ">
                <div style="
                    background: white;
                    border-radius: 12px;
                    padding: 20px;
                    border: 1px solid {WB_BORDER};
                    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
                ">
                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.05em;">Total Rows</div>
                    <div style="color: {WB_TEXT}; font-size: 1.8rem; font-weight: 700;">{len(chosen_df_raw):,}</div>
                </div>
                <div style="
                    background: white;
                    border-radius: 12px;
                    padding: 20px;
                    border: 1px solid {WB_BORDER};
                    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
                ">
                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.05em;">Columns</div>
                    <div style="color: {WB_TEXT}; font-size: 1.8rem; font-weight: 700;">{len(chosen_df_raw.columns)}</div>
                </div>
                <div style="
                    background: white;
                    border-radius: 12px;
                    padding: 20px;
                    border: 1px solid {WB_BORDER};
                    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
                ">
                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.05em;">DPD Threshold</div>
                    <div style="color: {WB_PRIMARY}; font-size: 1.8rem; font-weight: 700;">≥{dpd_threshold}</div>
                </div>
                <div style="
                    background: white;
                    border-radius: 12px;
                    padding: 20px;
                    border: 1px solid {WB_BORDER};
                    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
                ">
                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.05em;">Vintage Grouping</div>
                    <div style="color: {WB_PRIMARY}; font-size: 1.8rem; font-weight: 700;">{vintage_granularity_label}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

        # Vintage exclusion filter (applies to all downstream analyses)
        st.markdown(
            f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 16px 20px;
                border: 1px solid {WB_BORDER};
                margin-bottom: 16px;
            ">
                <div style="font-size: 0.85rem; font-weight: 600; color: {WB_TEXT}; margin-bottom: 4px;">
                    📅 Vintage Exclusion (Optional)
                </div>
                <div style="font-size: 0.78rem; color: {WB_MUTED};">
                    Drop specific vintage cohorts from the analysis (e.g. immature vintages
                    or a stress period). The exclusion applies to every tab — Data Integrity,
                    Summary Tables, Vintage Charts, and PD Optimizer.
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

        # Compute the vintage label for each loan using the currently selected granularity
        try:
            _od = pd.to_datetime(chosen_df_raw['Origination date'], errors='coerce')
            _g = (vintage_granularity or 'Q').upper()
            if _g == 'M':
                _vintage_series = _od.dt.to_period('M').astype(str)
            elif _g == 'H':
                _half = ((_od.dt.month - 1) // 6 + 1).astype('Int64')
                _vintage_series = (_od.dt.year.astype('Int64').astype(str)
                                   + 'H' + _half.astype(str))
            elif _g == 'Y':
                _vintage_series = _od.dt.year.astype('Int64').astype(str)
            else:
                _vintage_series = _od.dt.to_period('Q').astype(str)
            _all_vintages = sorted(
                v for v in _vintage_series.dropna().unique()
                if str(v) not in ('NaT', '<NA>', 'nan')
            )
        except Exception:
            _all_vintages = []
            _vintage_series = pd.Series(dtype='object')

        excluded_vintages = st.multiselect(
            '🗑️ Vintages to exclude',
            _all_vintages,
            default=[],
            help='Loans originated in the selected vintage cohorts will be removed '
                 'before any analysis. Vintage labels follow the "Vintage Grouping" '
                 'granularity selected above.',
            key='excluded_vintages',
        )

        if excluded_vintages and not _vintage_series.empty:
            keep_mask = ~_vintage_series.isin(excluded_vintages)
            chosen_df_raw = chosen_df_raw.loc[keep_mask]
            st.caption(
                f'📅 Excluding {len(excluded_vintages)} vintage(s): '
                f'{", ".join(str(v) for v in excluded_vintages)} — '
                f'{len(chosen_df_raw):,} rows remain.'
            )

        # Segmentation filter with card styling
        st.markdown(
            f"""
            <div style="
                background: white;
                border-radius: 12px;
                padding: 16px 20px;
                border: 1px solid {WB_BORDER};
                margin-bottom: 24px;
            ">
                <div style="font-size: 0.85rem; font-weight: 600; color: {WB_TEXT}; margin-bottom: 4px;">
                    🔍 Data Segmentation (Optional)
                </div>
                <div style="font-size: 0.78rem; color: {WB_MUTED};">
                    Apply up to 3 filters to narrow down your data. Filters are applied sequentially (AND logic). Numeric columns (e.g. Tenor) support a Range (≥ / ≤) mode.
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

        non_reserved = [c for c in chosen_df_raw.columns
                        if str(c).strip() not in RESERVED_COLS]

        # Initialise filter count in session state
        if 'num_filters' not in st.session_state:
            st.session_state['num_filters'] = 1

        # Controls to add / remove filters
        fctrl_cols = st.columns([1, 1, 3])
        with fctrl_cols[0]:
            if st.button('➕ Add filter', disabled=st.session_state['num_filters'] >= 3,
                         use_container_width=True):
                st.session_state['num_filters'] = min(3, st.session_state['num_filters'] + 1)
                st.rerun()
        with fctrl_cols[1]:
            if st.button('➖ Remove filter', disabled=st.session_state['num_filters'] <= 1,
                         use_container_width=True):
                st.session_state['num_filters'] = max(1, st.session_state['num_filters'] - 1)
                st.rerun()

        active_filter_cols = []  # track columns already used
        for filt_idx in range(st.session_state['num_filters']):
            available_cols = [c for c in non_reserved if c not in active_filter_cols]
            seg_col = st.selectbox(
                f'Filter {filt_idx + 1} — column',
                ['None'] + available_cols,
                key=f'seg_col_{filt_idx}',
            )
            if seg_col != 'None':
                active_filter_cols.append(seg_col)
                col_series = chosen_df_raw[seg_col]
                is_numeric = pd.api.types.is_numeric_dtype(col_series)

                if is_numeric:
                    mode = st.radio(
                        f'Filter mode for {seg_col}',
                        options=['Range (≥ / ≤)', 'Specific values'],
                        horizontal=True,
                        key=f'seg_mode_{filt_idx}',
                    )
                else:
                    mode = 'Specific values'

                if mode == 'Range (≥ / ≤)':
                    numeric_vals = pd.to_numeric(col_series, errors='coerce').dropna()
                    if numeric_vals.empty:
                        st.info(f'No numeric values found in {seg_col}.')
                    else:
                        raw_min = float(numeric_vals.min())
                        raw_max = float(numeric_vals.max())
                        # Use integer inputs when the column is integer-like to keep the UX clean
                        is_int_like = (pd.api.types.is_integer_dtype(col_series)
                                       or (raw_min.is_integer() and raw_max.is_integer()))
                        if is_int_like:
                            col_min, col_max, step = int(raw_min), int(raw_max), 1
                        else:
                            col_min, col_max, step = raw_min, raw_max, None
                        r1, r2 = st.columns(2)
                        with r1:
                            lo = st.number_input(
                                f'Min (≥) — {seg_col}',
                                value=col_min,
                                step=step,
                                key=f'seg_lo_{filt_idx}',
                            )
                        with r2:
                            hi = st.number_input(
                                f'Max (≤) — {seg_col}',
                                value=col_max,
                                step=step,
                                key=f'seg_hi_{filt_idx}',
                            )
                        if lo > hi:
                            st.warning(f'Min ({lo}) is greater than Max ({hi}); no rows will match.')
                        chosen_df_raw = chosen_df_raw[
                            (col_series >= lo) & (col_series <= hi)
                        ]
                else:
                    unique_vals = sorted(
                        [v for v in col_series.dropna().unique().tolist()],
                        key=lambda x: str(x)
                    )
                    selected_vals = st.multiselect(
                        f'Values for {seg_col}',
                        unique_vals,
                        default=unique_vals,
                        key=f'seg_vals_{filt_idx}',
                    )
                    if selected_vals:
                        chosen_df_raw = chosen_df_raw[chosen_df_raw[seg_col].isin(selected_vals)]
            if filt_idx < st.session_state['num_filters'] - 1:
                st.markdown(f"<hr style='margin:4px 0; border-color:{WB_BORDER};'>",
                            unsafe_allow_html=True)

        # Tabs with icons
        tab_integrity, tab_tables, tab_charts, tab_optimizer = st.tabs([
            "🛡️ Data Integrity",
            "📊 Summary Tables",
            "📈 Vintage Charts",
            "🎯 PD Optimizer",
        ])

        # ════════════════════════════════════════════════════════════════════
        # INTEGRITY TAB
        # ════════════════════════════════════════════════════════════════════
        with tab_integrity:
            st.markdown(
                f"""
                <div style="
                    background: linear-gradient(135deg, {WB_LIGHT} 0%, white 100%);
                    border-radius: 16px;
                    padding: 24px;
                    border: 1px solid {WB_BORDER};
                    margin-bottom: 24px;
                ">
                    <h3 style="margin: 0 0 8px 0; color: {WB_TEXT};">Data Quality Assessment</h3>
                    <p style="margin: 0; color: {WB_MUTED}; font-size: 0.9rem;">
                        Run comprehensive integrity checks to identify data quality issues, missing values, and inconsistencies in your loan portfolio data.
                    </p>
                </div>
                """,
                unsafe_allow_html=True
            )

            dataset_label = 'Full Dataset'

            col_btn, col_spacer = st.columns([1, 2])
            with col_btn:
                run_checks = st.button('🔍 Run Integrity Analysis', type='primary', use_container_width=True)

            if run_checks:
                with st.status('🔄 Analyzing data quality...', expanded=True) as status:
                    try:
                        st.write("Validating data schema...")
                        st.write("Checking date consistency...")
                        st.write("Analyzing value ranges...")
                        logger.info("Running integrity checks (%d rows)", len(chosen_df_raw))
                        _log_resource_usage("before integrity_checks")
                        summary, issues_df, vintage_issues_df, disappeared_df = run_integrity_checks(
                            chosen_df_raw, dpd_threshold=dpd_threshold,
                            vintage_granularity=vintage_granularity)
                        _log_resource_usage("after integrity_checks")
                        status.update(label='✅ Analysis complete!', state='complete')
                    except Exception:
                        logger.exception("CRASH during integrity checks")
                        status.update(label='❌ Analysis failed', state='error')
                        st.error(f"Integrity check error:\n```\n{traceback.format_exc()}\n```")
                        summary = {'fatal': 'Unexpected error — see details above'}
                        issues_df = vintage_issues_df = disappeared_df = pd.DataFrame()

                if 'fatal' in summary:
                    st.error(f"❌ Critical Error: {summary['fatal']}")
                else:
                    # Store flagged row indices for optional exclusion in curves
                    if issues_df is not None and not issues_df.empty and 'index' in issues_df.columns:
                        st.session_state['flagged_indices'] = set(issues_df['index'].tolist())
                    else:
                        st.session_state['flagged_indices'] = set()

                    st.success('✅ Integrity checks completed successfully!')

                    # Results summary card
                    st.markdown(
                        f"""
                        <div style="
                            background: white;
                            border-radius: 12px;
                            padding: 20px;
                            border: 1px solid {WB_BORDER};
                            margin: 16px 0;
                        ">
                            <h4 style="margin: 0 0 16px 0; color: {WB_TEXT};">📋 Analysis Summary</h4>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                    st.json(summary)

                    # Download buttons in columns
                    st.markdown("#### 📥 Export Reports")
                    dl_col1, dl_col2, dl_col3 = st.columns(3)

                    with dl_col1:
                        pdf_bytes = export_integrity_pdf(summary, dataset_label=dataset_label,
                                                         vintage_issues_df=vintage_issues_df)
                        st.download_button(
                            '📄 PDF Report',
                            pdf_bytes,
                            'integrity_report.pdf',
                            'application/pdf',
                            use_container_width=True
                        )

                    with dl_col2:
                        xlsx_bytes = export_issues_excel(issues_df, vintage_issues_df)
                        st.download_button(
                            '📊 Excel Issues',
                            xlsx_bytes,
                            'integrity_issues_sample.xlsx',
                            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width=True
                        )

                    with dl_col3:
                        if disappeared_df is not None and not disappeared_df.empty:
                            dis_bytes = export_disappeared_loans_excel(disappeared_df)
                            st.download_button(
                                '🔍 Disappeared Loans',
                                dis_bytes,
                                'disappeared_loans.xlsx',
                                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                use_container_width=True
                            )

                    st.markdown("---")

                    # Issues display
                    if issues_df is not None and not issues_df.empty:
                        st.markdown(f"#### ⚠️ Row-Level Issues ({len(issues_df)} samples)")
                        st.dataframe(issues_df.head(500), use_container_width=True, height=300)
                    else:
                        st.info('✅ No row-level data quality issues detected.')

                    if vintage_issues_df is not None and not vintage_issues_df.empty:
                        st.markdown(f"#### 📊 Vintage/Cohort Issues ({len(vintage_issues_df)} items)")
                        st.dataframe(vintage_issues_df.head(500), use_container_width=True, height=300)
                    else:
                        st.info('✅ No vintage-level issues detected.')

        # ════════════════════════════════════════════════════════════════════
        # TABLES TAB
        # ════════════════════════════════════════════════════════════════════
        with tab_tables:
            subtab_vintage, subtab_segment = st.tabs([
                "📊 Vintage Analysis",
                "🔎 PD Segmentation Analysis",
            ])

            # ── Sub-tab: Vintage Analysis (original) ──────────────────────
            with subtab_vintage:
                st.markdown(
                    f"""
                    <div style="
                        background: linear-gradient(135deg, {WB_LIGHT} 0%, white 100%);
                        border-radius: 16px;
                        padding: 24px;
                        border: 1px solid {WB_BORDER};
                        margin-bottom: 24px;
                    ">
                        <h3 style="margin: 0 0 8px 0; color: {WB_TEXT};">Vintage Performance Summary</h3>
                        <p style="margin: 0; color: {WB_MUTED}; font-size: 0.9rem;">
                            Comprehensive breakdown of loan performance metrics by vintage cohort, including cumulative default rates and annualized statistics.
                        </p>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                pd_calc_method_table = st.selectbox(
                    '📐 PD Calculation Method',
                    ['By Loan Count', 'By Amount (Current / Origination)'],
                    index=0,
                    help='By Loan Count: defaulted loans / total loans. '
                         'By Amount: sum of Current Amount at default date / sum of Origination Amount.',
                    key='pd_calc_method_table',
                )
                pd_by_amount_table = pd_calc_method_table == 'By Amount (Current / Origination)'

                try:
                    summary_df = compute_vintage_default_summary(
                        chosen_df_raw, dpd_threshold=dpd_threshold,
                        pd_by_amount=pd_by_amount_table,
                        vintage_granularity=vintage_granularity)

                    # Key metrics cards
                    total_loans = summary_df['Unique_loans'].sum()
                    total_defaults = summary_df['Defaulted_loans'].sum()
                    if pd_by_amount_table:
                        total_orig_amt = summary_df['Total_Origination_Amount'].sum()
                        total_def_amt = summary_df['Total_Default_Amount'].sum()
                        avg_pd = (total_def_amt / total_orig_amt * 100) if total_orig_amt > 0 else 0
                        weights = summary_df['Total_Origination_Amount']
                    else:
                        avg_pd = (total_defaults / total_loans * 100) if total_loans > 0 else 0
                        weights = summary_df['Unique_loans']

                    # Annualized PD average across all rows: derive from aggregate Cum_PD
                    # and weighted observation time so it stays consistent with Avg Default Rate.
                    obs_times = summary_df['Observation_Time']
                    obs_mask = obs_times.notna() & (obs_times > 0) & (weights > 0)
                    if obs_mask.any() and weights[obs_mask].sum() > 0:
                        weighted_obs_time = (obs_times[obs_mask] * weights[obs_mask]).sum() / weights[obs_mask].sum()
                        cum_pd_frac = avg_pd / 100.0
                        if weighted_obs_time > 0 and cum_pd_frac < 1:
                            avg_annualized_pd = (1 - (1 - cum_pd_frac) ** (1 / weighted_obs_time)) * 100
                        else:
                            avg_annualized_pd = avg_pd
                    else:
                        avg_annualized_pd = 0

                    st.markdown(
                        f"""
                        <div style="
                            display: grid;
                            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                            gap: 16px;
                            margin-bottom: 24px;
                        ">
                            <div style="
                                background: white;
                                border-radius: 12px;
                                padding: 20px;
                                border: 1px solid {WB_BORDER};
                                text-align: center;
                            ">
                                <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Total Loans</div>
                                <div style="color: {WB_TEXT}; font-size: 1.5rem; font-weight: 700;">{total_loans:,}</div>
                            </div>
                            <div style="
                                background: white;
                                border-radius: 12px;
                                padding: 20px;
                                border: 1px solid {WB_BORDER};
                                text-align: center;
                            ">
                                <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Total Defaults</div>
                                <div style="color: #EF4444; font-size: 1.5rem; font-weight: 700;">{total_defaults:,}</div>
                            </div>
                            <div style="
                                background: white;
                                border-radius: 12px;
                                padding: 20px;
                                border: 1px solid {WB_BORDER};
                                text-align: center;
                            ">
                                <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Avg Default Rate {'(by Amt)' if pd_by_amount_table else '(by Count)'}</div>
                                <div style="color: {WB_PRIMARY}; font-size: 1.5rem; font-weight: 700;">{avg_pd:.2f}%</div>
                            </div>
                            <div style="
                                background: white;
                                border-radius: 12px;
                                padding: 20px;
                                border: 1px solid {WB_BORDER};
                                text-align: center;
                            ">
                                <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Annualized PD Avg</div>
                                <div style="color: {WB_PRIMARY}; font-size: 1.5rem; font-weight: 700;">{avg_annualized_pd:.2f}%</div>
                            </div>
                            <div style="
                                background: white;
                                border-radius: 12px;
                                padding: 20px;
                                border: 1px solid {WB_BORDER};
                                text-align: center;
                            ">
                                <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Vintages</div>
                                <div style="color: {WB_TEXT}; font-size: 1.5rem; font-weight: 700;">{len(summary_df)}</div>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                    st.markdown(
                        f"""
                        <div style="
                            background: white;
                            border-radius: 12px;
                            padding: 16px 20px;
                            border: 1px solid {WB_BORDER};
                            margin-bottom: 16px;
                        ">
                            <div style="font-size: 0.9rem; font-weight: 600; color: {WB_TEXT};">📊 Vintage Default Summary Table {'(by Amount)' if pd_by_amount_table else '(by Loan Count)'}</div>
                            <div style="font-size: 0.8rem; color: {WB_MUTED};">
                                {'Cum PD = Σ Current Amount at default / Σ Origination Amount. ' if pd_by_amount_table else ''}Observation Time = default date − first observation (if defaulted), else last observation − first observation (in years)
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                    # Rename only for display
                    rename_map = {
                        "Unique_loans": "Unique loans",
                        "Defaulted_loans": "Defaulted loans",
                        "Observation_Time": "Obs Time (years)",
                        "Default_rate_pa": "Annualized default rate",
                        "Cum_PD": "Cum PD",
                    }
                    if pd_by_amount_table:
                        rename_map["Total_Origination_Amount"] = "Total Origination Amt"
                        rename_map["Total_Default_Amount"] = "Default Amt (Current)"
                    disp = summary_df.rename(columns=rename_map)
                    disp["Cum PD (%)"] = disp["Cum PD"] * 100
                    disp["Annualized default rate (%)"] = disp["Annualized default rate"] * 100

                    if pd_by_amount_table:
                        table = disp[[
                            "Vintage",
                            "Unique loans",
                            "Defaulted loans",
                            "Total Origination Amt",
                            "Default Amt (Current)",
                            "Cum PD (%)",
                            "Obs Time (years)",
                            "Annualized default rate (%)",
                        ]]
                        styles = {
                            "Unique loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                            "Defaulted loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                            "Total Origination Amt": "{:,.2f}" if pretty_ints else "{:.2f}",
                            "Default Amt (Current)": "{:,.2f}" if pretty_ints else "{:.2f}",
                            "Cum PD (%)": "{:.2f}",
                            "Obs Time (years)": "{:.2f}",
                            "Annualized default rate (%)": "{:.2f}",
                        }
                    else:
                        table = disp[[
                            "Vintage",
                            "Unique loans",
                            "Defaulted loans",
                            "Cum PD (%)",
                            "Obs Time (years)",
                            "Annualized default rate (%)",
                        ]]
                        styles = {
                            "Unique loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                            "Defaulted loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                            "Cum PD (%)": "{:.2f}",
                            "Obs Time (years)": "{:.2f}",
                            "Annualized default rate (%)": "{:.2f}",
                        }
                    styler = (
                        table.style
                        .format(styles)
                        .background_gradient(subset=["Cum PD (%)", "Annualized default rate (%)"], cmap="Blues")
                        .hide(axis="index")
                    )

                    st.dataframe(styler, use_container_width=True, height=400)

                    st.markdown("#### 📥 Export Data")
                    st.download_button(
                        '📊 Download CSV',
                        summary_df.to_csv(index=False).encode('utf-8'),
                        'vintage_default_summary.csv',
                        'text/csv',
                        use_container_width=False
                    )
                except Exception as e:
                    logger.exception("CRASH in vintage default summary")
                    st.error(f'❌ Could not compute vintage default summary: {e}')

            # ── Sub-tab: PD Segmentation Analysis ─────────────────────────
            with subtab_segment:
                st.markdown(
                    f"""
                    <div style="
                        background: linear-gradient(135deg, {WB_LIGHT} 0%, white 100%);
                        border-radius: 16px;
                        padding: 24px;
                        border: 1px solid {WB_BORDER};
                        margin-bottom: 24px;
                    ">
                        <h3 style="margin: 0 0 8px 0; color: {WB_TEXT};">PD Segmentation Analysis</h3>
                        <p style="margin: 0; color: {WB_MUTED}; font-size: 0.9rem;">
                            Analyse default rates grouped by any data attribute (e.g. Tenor, Product, Region).
                            Each row represents a distinct segment value instead of a vintage cohort.
                        </p>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                seg_non_reserved = [c for c in chosen_df_raw.columns
                                    if str(c).strip() not in RESERVED_COLS]

                if not seg_non_reserved:
                    st.warning('No non-reserved columns available for segmentation. '
                               'Your dataset only contains the core loan columns.')
                else:
                    seg_col_choice = st.selectbox(
                        '🏷️ Segmentation Column',
                        seg_non_reserved,
                        index=0,
                        help='Choose the column whose distinct values will become the rows of the PD table.',
                        key='seg_col_pd_analysis',
                    )

                    pd_calc_method_seg = st.selectbox(
                        '📐 PD Calculation Method',
                        ['By Loan Count', 'By Amount (Current / Origination)'],
                        index=0,
                        help='By Loan Count: defaulted loans / total loans. '
                             'By Amount: sum of Current Amount at default date / sum of Origination Amount.',
                        key='pd_calc_method_seg',
                    )
                    pd_by_amount_seg = pd_calc_method_seg == 'By Amount (Current / Origination)'

                    try:
                        seg_summary_df = compute_segment_default_summary(
                            chosen_df_raw,
                            dpd_threshold=dpd_threshold,
                            segment_col=seg_col_choice,
                            pd_by_amount=pd_by_amount_seg,
                            vintage_granularity=vintage_granularity,
                        )

                        # Key metrics cards
                        seg_total_loans = seg_summary_df['Unique_loans'].sum()
                        seg_total_defaults = seg_summary_df['Defaulted_loans'].sum()
                        if pd_by_amount_seg:
                            seg_total_orig = seg_summary_df['Total_Origination_Amount'].sum()
                            seg_total_def = seg_summary_df['Total_Default_Amount'].sum()
                            seg_avg_pd = (seg_total_def / seg_total_orig * 100) if seg_total_orig > 0 else 0
                            seg_weights = seg_summary_df['Total_Origination_Amount']
                        else:
                            seg_avg_pd = (seg_total_defaults / seg_total_loans * 100) if seg_total_loans > 0 else 0
                            seg_weights = seg_summary_df['Unique_loans']

                        # Annualized PD average across all rows
                        seg_obs_times = seg_summary_df['Observation_Time']
                        seg_obs_mask = seg_obs_times.notna() & (seg_obs_times > 0) & (seg_weights > 0)
                        if seg_obs_mask.any() and seg_weights[seg_obs_mask].sum() > 0:
                            seg_weighted_obs_time = (seg_obs_times[seg_obs_mask] * seg_weights[seg_obs_mask]).sum() / seg_weights[seg_obs_mask].sum()
                            seg_cum_pd_frac = seg_avg_pd / 100.0
                            if seg_weighted_obs_time > 0 and seg_cum_pd_frac < 1:
                                seg_avg_annualized_pd = (1 - (1 - seg_cum_pd_frac) ** (1 / seg_weighted_obs_time)) * 100
                            else:
                                seg_avg_annualized_pd = seg_avg_pd
                        else:
                            seg_avg_annualized_pd = 0

                        st.markdown(
                            f"""
                            <div style="
                                display: grid;
                                grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                                gap: 16px;
                                margin-bottom: 24px;
                            ">
                                <div style="
                                    background: white;
                                    border-radius: 12px;
                                    padding: 20px;
                                    border: 1px solid {WB_BORDER};
                                    text-align: center;
                                ">
                                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Total Loans</div>
                                    <div style="color: {WB_TEXT}; font-size: 1.5rem; font-weight: 700;">{seg_total_loans:,}</div>
                                </div>
                                <div style="
                                    background: white;
                                    border-radius: 12px;
                                    padding: 20px;
                                    border: 1px solid {WB_BORDER};
                                    text-align: center;
                                ">
                                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Total Defaults</div>
                                    <div style="color: #EF4444; font-size: 1.5rem; font-weight: 700;">{seg_total_defaults:,}</div>
                                </div>
                                <div style="
                                    background: white;
                                    border-radius: 12px;
                                    padding: 20px;
                                    border: 1px solid {WB_BORDER};
                                    text-align: center;
                                ">
                                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Avg Default Rate {'(by Amt)' if pd_by_amount_seg else '(by Count)'}</div>
                                    <div style="color: {WB_PRIMARY}; font-size: 1.5rem; font-weight: 700;">{seg_avg_pd:.2f}%</div>
                                </div>
                                <div style="
                                    background: white;
                                    border-radius: 12px;
                                    padding: 20px;
                                    border: 1px solid {WB_BORDER};
                                    text-align: center;
                                ">
                                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Annualized PD Avg</div>
                                    <div style="color: {WB_PRIMARY}; font-size: 1.5rem; font-weight: 700;">{seg_avg_annualized_pd:.2f}%</div>
                                </div>
                                <div style="
                                    background: white;
                                    border-radius: 12px;
                                    padding: 20px;
                                    border: 1px solid {WB_BORDER};
                                    text-align: center;
                                ">
                                    <div style="color: {WB_MUTED}; font-size: 0.75rem; text-transform: uppercase;">Segments</div>
                                    <div style="color: {WB_TEXT}; font-size: 1.5rem; font-weight: 700;">{len(seg_summary_df)}</div>
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

                        st.markdown(
                            f"""
                            <div style="
                                background: white;
                                border-radius: 12px;
                                padding: 16px 20px;
                                border: 1px solid {WB_BORDER};
                                margin-bottom: 16px;
                            ">
                                <div style="font-size: 0.9rem; font-weight: 600; color: {WB_TEXT};">🔎 PD by {seg_col_choice} {'(by Amount)' if pd_by_amount_seg else '(by Loan Count)'}</div>
                                <div style="font-size: 0.8rem; color: {WB_MUTED};">
                                    {'Cum PD = Σ Current Amount at default / Σ Origination Amount. ' if pd_by_amount_seg else ''}Observation Time = default date − first observation (if defaulted), else last observation − first observation (in years)
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

                        # Rename for display
                        seg_rename = {
                            "Segment": seg_col_choice,
                            "Unique_loans": "Unique loans",
                            "Defaulted_loans": "Defaulted loans",
                            "Observation_Time": "Obs Time (years)",
                            "Default_rate_pa": "Annualized default rate",
                            "Cum_PD": "Cum PD",
                        }
                        if pd_by_amount_seg:
                            seg_rename["Total_Origination_Amount"] = "Total Origination Amt"
                            seg_rename["Total_Default_Amount"] = "Default Amt (Current)"
                        seg_disp = seg_summary_df.rename(columns=seg_rename)
                        seg_disp["Cum PD (%)"] = seg_disp["Cum PD"] * 100
                        seg_disp["Annualized default rate (%)"] = seg_disp["Annualized default rate"] * 100

                        if pd_by_amount_seg:
                            seg_table = seg_disp[[
                                seg_col_choice,
                                "Unique loans",
                                "Defaulted loans",
                                "Total Origination Amt",
                                "Default Amt (Current)",
                                "Cum PD (%)",
                                "Obs Time (years)",
                                "Annualized default rate (%)",
                            ]]
                            seg_styles = {
                                "Unique loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                                "Defaulted loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                                "Total Origination Amt": "{:,.2f}" if pretty_ints else "{:.2f}",
                                "Default Amt (Current)": "{:,.2f}" if pretty_ints else "{:.2f}",
                                "Cum PD (%)": "{:.2f}",
                                "Obs Time (years)": "{:.2f}",
                                "Annualized default rate (%)": "{:.2f}",
                            }
                        else:
                            seg_table = seg_disp[[
                                seg_col_choice,
                                "Unique loans",
                                "Defaulted loans",
                                "Cum PD (%)",
                                "Obs Time (years)",
                                "Annualized default rate (%)",
                            ]]
                            seg_styles = {
                                "Unique loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                                "Defaulted loans": "{:,.0f}" if pretty_ints else "{:.0f}",
                                "Cum PD (%)": "{:.2f}",
                                "Obs Time (years)": "{:.2f}",
                                "Annualized default rate (%)": "{:.2f}",
                            }

                        seg_styler = (
                            seg_table.style
                            .format(seg_styles)
                            .background_gradient(subset=["Cum PD (%)", "Annualized default rate (%)"], cmap="Blues")
                            .hide(axis="index")
                        )

                        st.dataframe(seg_styler, use_container_width=True, height=400)

                        st.markdown("#### 📥 Export Data")
                        st.download_button(
                            '📊 Download CSV',
                            seg_summary_df.to_csv(index=False).encode('utf-8'),
                            f'pd_by_{seg_col_choice}.csv',
                            'text/csv',
                            use_container_width=False,
                            key='seg_csv_download',
                        )
                    except Exception as e:
                        logger.exception("CRASH in segment PD summary")
                        st.error(f'❌ Could not compute segment PD summary: {e}')

        # ════════════════════════════════════════════════════════════════════
        # CHARTS TAB
        # ════════════════════════════════════════════════════════════════════
        with tab_charts:
            st.markdown(
                f"""
                <div style="
                    background: linear-gradient(135deg, {WB_LIGHT} 0%, white 100%);
                    border-radius: 16px;
                    padding: 24px;
                    border: 1px solid {WB_BORDER};
                    margin-bottom: 24px;
                ">
                    <h3 style="margin: 0 0 8px 0; color: {WB_TEXT};">Vintage Curve Visualization</h3>
                    <p style="margin: 0; color: {WB_MUTED}; font-size: 0.9rem;">
                        Interactive default rate curves showing cumulative performance evolution across loan cohorts over time.
                    </p>
                </div>
                """,
                unsafe_allow_html=True
            )

            col_chart, col_settings = st.columns([3, 1], gap="large")

            with col_settings:
                # Settings panel card
                st.markdown(
                    f"""
                    <div style="
                        background: white;
                        border-radius: 12px;
                        padding: 20px;
                        border: 1px solid {WB_BORDER};
                        margin-bottom: 16px;
                    ">
                        <div style="display: flex; align-items: center; gap: 10px; margin-bottom: 16px;">
                            <div style="
                                background: linear-gradient(135deg, {WB_PRIMARY} 0%, {WB_SECONDARY} 100%);
                                width: 32px;
                                height: 32px;
                                border-radius: 8px;
                                display: flex;
                                align-items: center;
                                justify-content: center;
                            ">
                                <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2">
                                    <line x1="4" y1="21" x2="4" y2="14"></line>
                                    <line x1="4" y1="10" x2="4" y2="3"></line>
                                    <line x1="12" y1="21" x2="12" y2="12"></line>
                                    <line x1="12" y1="8" x2="12" y2="3"></line>
                                    <line x1="20" y1="21" x2="20" y2="16"></line>
                                    <line x1="20" y1="12" x2="20" y2="3"></line>
                                    <line x1="1" y1="14" x2="7" y2="14"></line>
                                    <line x1="9" y1="8" x2="15" y2="8"></line>
                                    <line x1="17" y1="16" x2="23" y2="16"></line>
                                </svg>
                            </div>
                            <span style="font-weight: 600; color: {WB_TEXT};">Chart Settings</span>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                granularity = st.selectbox('📅 Time Granularity', ['Quarterly (QOB)', 'Monthly (MOB)'], index=0)
                gran_key = 'QOB' if 'QOB' in granularity else 'MOB'
                max_months_show = st.slider('📏 Max Months to Display', min_value=12, max_value=180, value=60, step=6)

                st.markdown("---")
                st.markdown("**PD Calculation**")
                pd_calc_method_chart = st.selectbox(
                    '📐 PD Method',
                    ['By Loan Count', 'By Amount (Current / Origination)'],
                    index=0,
                    help='By Loan Count: defaulted loans / total loans. '
                         'By Amount: sum of Current Amount at default date / sum of Origination Amount.',
                    key='pd_calc_method_chart',
                )
                pd_by_amount_chart = pd_calc_method_chart == 'By Amount (Current / Origination)'

                st.markdown("---")
                st.markdown("**Curve Processing**")
                smooth_curves = st.checkbox('🔄 Smooth curves', value=True)
                force_monotone = st.checkbox('📈 Force monotone (cummax)', value=True)
                cure_adjusted = st.checkbox(
                    '💊 Cure-adjusted mode',
                    value=False,
                    help='Use current DPD status at each snapshot instead of cumulative ever-defaulted flag.'
                )
                exclude_flagged = st.checkbox(
                    '🚫 Exclude flagged rows',
                    value=False,
                    help='Exclude rows flagged by integrity checks. Run integrity checks first.'
                )

                st.markdown("---")
                st.markdown("**Visual Options**")
                show_legend = st.checkbox('📋 Show legend', value=True)
                palette_option = st.selectbox('🎨 Color palette', ['Gradient', 'Plotly', 'Viridis'])
                base_color = st.color_picker(
                    '🖌️ Base color',
                    value=WB_PRIMARY,
                    help='Used when Gradient palette is selected.'
                )
                line_width = st.slider('✏️ Line width', min_value=1, max_value=5, value=2)
                line_style = st.selectbox(
                    '〰️ Line style',
                    ['solid', 'dash', 'dot', 'dashdot'],
                    index=0,
                )
                show_markers = st.checkbox('🔘 Show data-point markers', value=False)

                st.markdown("---")
                st.markdown("**Layout & Axes**")
                chart_title = st.text_input(
                    '📝 Custom chart title',
                    value='',
                    help='Leave blank to use the auto-generated title.',
                )
                chart_height = st.slider('📐 Chart height (px)', min_value=350, max_value=900, value=550, step=50)
                title_font_size = st.slider('🔤 Title font size', min_value=12, max_value=28, value=18, step=1)
                axis_font_size = st.slider('🔡 Axis label size', min_value=9, max_value=18, value=13, step=1)
                show_grid = st.checkbox('🔲 Show gridlines', value=True)
                y_axis_max = st.number_input(
                    '📊 Y-axis max (%)',
                    min_value=0.0,
                    max_value=100.0,
                    value=0.0,
                    step=1.0,
                    help='Set to 0 for auto-scale.',
                )
                bg_color = st.color_picker('🎨 Plot background', value='#FFFFFF')

            with col_chart:
                try:
                    prog_bar = st.progress(0.0, text="Initializing chart pipeline...")
                    upd = mk_progress_updater(prog_bar, steps=5)

                    if gran_key == 'QOB':
                        max_periods = max(1, math.ceil(max_months_show / 3))
                    else:
                        max_periods = max_months_show

                    excl = st.session_state.get('flagged_indices', set()) if exclude_flagged else None

                    df_plot_any, cohort_sizes = build_chart_data_fast(
                        chosen_df_raw, dpd_threshold=dpd_threshold,
                        max_periods=max_periods,
                        granularity=gran_key,
                        smooth=smooth_curves,
                        force_monotone=force_monotone,
                        cure_adjusted=cure_adjusted,
                        exclude_indices=excl,
                        pd_by_amount=pd_by_amount_chart,
                        vintage_granularity=vintage_granularity,
                        prog=upd,
                    )

                    if df_plot_any.empty:
                        prog_bar.progress(1.0, text="Complete")
                        st.warning('⚠️ Insufficient data to generate curves for the selected parameters.')
                    else:
                        vintages = df_plot_any.columns.tolist()
                        st.markdown(
                            f"""
                            <div style="
                                background: white;
                                border-radius: 10px;
                                padding: 12px 16px;
                                border: 1px solid {WB_BORDER};
                                margin-bottom: 16px;
                            ">
                                <span style="font-size: 0.85rem; font-weight: 500; color: {WB_TEXT};">
                                    🏷️ Select vintages to display ({len(vintages)} available)
                                </span>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                        selected_vintages = st.multiselect(
                            'Select vintages',
                            vintages,
                            default=vintages,
                            label_visibility="collapsed"
                        )

                        if not selected_vintages:
                            prog_bar.progress(1.0, text="Complete")
                            st.info('ℹ️ Select at least one vintage to display the chart.')
                        else:
                            df_plot = df_plot_any[selected_vintages]
                            prog_bar.progress(0.9, text="Rendering visualization...")
                            if chart_title.strip():
                                ttl = chart_title.strip()
                            else:
                                method_label = 'by Amount' if pd_by_amount_chart else 'by Count'
                                ttl = f'Vintage Default-Rate Evolution ({method_label}) | DPD ≥ {dpd_threshold}'
                            if not chart_title.strip() and cure_adjusted:
                                ttl += ' | Cure-Adjusted'
                            fig = plot_curves_percent_with_months(
                                df_wide=df_plot,
                                title=ttl,
                                show_legend=show_legend,
                                legend_limit=50,
                                palette=palette_option,
                                base_color=base_color,
                                line_width=line_width,
                                cohort_sizes=cohort_sizes,
                                line_style=line_style,
                                show_markers=show_markers,
                                chart_height=chart_height,
                                title_font_size=title_font_size,
                                axis_font_size=axis_font_size,
                                show_grid=show_grid,
                                y_axis_max=y_axis_max,
                                bg_color=bg_color,
                                pd_by_amount=pd_by_amount_chart,
                            )
                            prog_bar.progress(1.0, text="✅ Chart ready")

                            st.plotly_chart(fig, use_container_width=True, config={
                                'displayModeBar': True,
                                'displaylogo': False,
                                'modeBarButtonsToRemove': ['lasso2d', 'select2d']
                            })

                            # Export section
                            st.markdown("#### 📥 Export Chart Data")
                            period_col = df_plot.index.name or gran_key
                            export_df = df_plot.reset_index()
                            if gran_key == 'QOB':
                                export_df.insert(0, "Months", export_df[period_col] * 3)

                            exp_col1, exp_col2, exp_col3 = st.columns([1, 1, 2])
                            with exp_col1:
                                st.download_button(
                                    f'📊 Download CSV',
                                    export_df.to_csv(index=False).encode('utf-8'),
                                    f'vintage_curves_{gran_key.lower()}.csv',
                                    'text/csv',
                                    use_container_width=True
                                )
                            with exp_col2:
                                try:
                                    _summary = compute_vintage_default_summary(
                                        chosen_df_raw, dpd_threshold=dpd_threshold,
                                        pd_by_amount=pd_by_amount_chart,
                                        vintage_granularity=vintage_granularity)
                                    xlsx_bytes = export_consistency_excel(
                                        _summary, df_plot, gran_key)
                                    st.download_button(
                                        '📋 Consistency Check (Excel)',
                                        xlsx_bytes,
                                        'table_vs_chart_consistency.xlsx',
                                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                        use_container_width=True
                                    )
                                except Exception:
                                    pass

                except Exception as e:
                    logger.exception("CRASH in chart generation")
                    st.error(f'❌ Chart generation failed: {e}')

        # ════════════════════════════════════════════════════════════════════
        # PD OPTIMIZER TAB
        # ════════════════════════════════════════════════════════════════════
        with tab_optimizer:
            st.markdown(
                f"""
                <div style="
                    background: linear-gradient(135deg, {WB_LIGHT} 0%, white 100%);
                    border-radius: 16px;
                    padding: 24px;
                    border: 1px solid {WB_BORDER};
                    margin-bottom: 24px;
                ">
                    <h3 style="margin: 0 0 8px 0; color: {WB_TEXT};">🎯 Target PD Optimizer</h3>
                    <p style="margin: 0; color: {WB_MUTED}; font-size: 0.9rem;">
                        Enter the annualized PD you want to hit and the tool scans every non-reserved
                        attribute to recommend single-filter eligibility rules (with the highest retention)
                        that bring the portfolio at or below target. A two-filter combo is also explored.
                    </p>
                </div>
                """,
                unsafe_allow_html=True
            )

            opt_c1, opt_c2, opt_c3 = st.columns([1.2, 1.2, 1])
            with opt_c1:
                target_pd_pct = st.number_input(
                    '🎯 Target annualized PD (%)',
                    min_value=0.0, max_value=100.0,
                    value=5.0, step=0.1,
                    help='Desired per-annum default rate for the filtered portfolio.',
                    key='opt_target_pd',
                )
            with opt_c2:
                pd_calc_method_opt = st.selectbox(
                    '📐 PD Calculation Method',
                    ['By Loan Count', 'By Amount (Current / Origination)'],
                    index=0,
                    help='Retention is ranked by the same basis — loans or origination amount.',
                    key='pd_calc_method_opt',
                )
                pd_by_amount_opt = pd_calc_method_opt == 'By Amount (Current / Origination)'
            with opt_c3:
                min_retention_pct = st.slider(
                    'Min retention (%)',
                    min_value=1, max_value=80, value=5, step=1,
                    help='Discard filters that keep fewer than this share of the baseline portfolio.',
                    key='opt_min_retention',
                )

            # Constraints — lock columns / force required values
            non_reserved_for_opt = [c for c in chosen_df_raw.columns
                                    if str(c).strip() not in RESERVED_COLS]
            locked_cols = st.multiselect(
                '🔒 Lock columns (no filter will be proposed)',
                non_reserved_for_opt,
                default=[],
                help='Attributes listed here are kept as-is — the optimizer will not '
                     'suggest any filter on them. Example: pick "Industry" to keep all '
                     'industries. Combine with the sidebar filters to pre-narrow the data.',
                key='opt_locked_cols',
            )

            required_map: dict = {}
            with st.expander('🔧 Advanced: force required values on specific columns', expanded=False):
                st.caption(
                    'Pick columns that can still be filtered, but only in ways that '
                    'preserve the values / range you list here. Any suggestion that '
                    'would drop a required value/range is rejected.'
                )
                constrainable = [c for c in non_reserved_for_opt if c not in locked_cols]
                constrained_cols = st.multiselect(
                    'Columns with required values/ranges',
                    constrainable,
                    default=[],
                    key='opt_required_cols',
                )
                for c in constrained_cols:
                    col_s = chosen_df_raw[c]
                    if pd.api.types.is_numeric_dtype(col_s):
                        vals = pd.to_numeric(col_s, errors='coerce').dropna()
                        if vals.empty:
                            st.info(f'No numeric values in {c}.')
                            continue
                        raw_min, raw_max = float(vals.min()), float(vals.max())
                        is_int_like = (pd.api.types.is_integer_dtype(col_s)
                                       or (raw_min.is_integer() and raw_max.is_integer()))
                        if is_int_like:
                            step, vmin, vmax = 1, int(raw_min), int(raw_max)
                        else:
                            step, vmin, vmax = None, raw_min, raw_max
                        rc1, rc2 = st.columns(2)
                        with rc1:
                            lo_req = st.number_input(
                                f'{c} — required Min',
                                value=vmin, step=step,
                                key=f'opt_req_lo_{c}',
                            )
                        with rc2:
                            hi_req = st.number_input(
                                f'{c} — required Max',
                                value=vmax, step=step,
                                key=f'opt_req_hi_{c}',
                            )
                        required_map[c] = {'type': 'numeric', 'lo': float(lo_req), 'hi': float(hi_req)}
                    else:
                        unique_vals = sorted(
                            [v for v in col_s.dropna().unique().tolist()],
                            key=lambda x: str(x),
                        )
                        sel = st.multiselect(
                            f'{c} — required values (must remain in any suggestion)',
                            unique_vals,
                            default=unique_vals,
                            key=f'opt_req_vals_{c}',
                        )
                        required_map[c] = {'type': 'categorical', 'values': sel}

            run_opt = st.button('🚀 Generate recommendations', type='primary', key='opt_run')

            if run_opt:
                with st.status('🔍 Scanning attributes for best eligibility rules…',
                               expanded=False) as status:
                    try:
                        loan_tbl = compute_loan_level_table(
                            chosen_df_raw,
                            dpd_threshold=dpd_threshold,
                            vintage_granularity=vintage_granularity,
                        )
                        candidate_cols = [c for c in loan_tbl.columns
                                          if c not in _LOAN_INTERNAL_COLS
                                          and not str(c).startswith('__')
                                          and c not in locked_cols]
                        feasible_df, partial_df, baseline = suggest_filters_for_target_pd(
                            loan_tbl,
                            target_ann_pd=target_pd_pct / 100.0,
                            candidate_cols=candidate_cols,
                            pd_by_amount=pd_by_amount_opt,
                            min_retention=min_retention_pct / 100.0,
                            required_map=required_map,
                        )
                        combo = suggest_combo_filter(
                            loan_tbl,
                            target_ann_pd=target_pd_pct / 100.0,
                            feasible_df=feasible_df,
                            partial_df=partial_df,
                            pd_by_amount=pd_by_amount_opt,
                        )
                        status.update(label='✅ Scan complete.', state='complete')
                    except Exception as e:
                        logger.exception("CRASH in PD optimizer")
                        status.update(label=f'❌ Optimizer failed: {e}', state='error')
                        feasible_df = partial_df = pd.DataFrame()
                        baseline = (np.nan, np.nan, 0, 0.0)
                        combo = None

                base_cp, base_ap, base_n, base_amt = baseline
                base_ap_pct = base_ap * 100 if base_ap is not None and not np.isnan(base_ap) else float('nan')
                base_cp_pct = base_cp * 100 if base_cp is not None and not np.isnan(base_cp) else float('nan')
                gap_pct = base_ap_pct - target_pd_pct if not np.isnan(base_ap_pct) else float('nan')
                on_target = not np.isnan(base_ap_pct) and base_ap_pct <= target_pd_pct

                # Baseline summary card
                st.markdown(
                    f"""
                    <div style="
                        display: grid;
                        grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
                        gap: 16px;
                        margin-bottom: 20px;
                    ">
                        <div style="background: white; border-radius: 12px; padding: 18px;
                             border: 1px solid {WB_BORDER}; text-align: center;">
                            <div style="color: {WB_MUTED}; font-size: 0.72rem; text-transform: uppercase;">Baseline Loans</div>
                            <div style="color: {WB_TEXT}; font-size: 1.5rem; font-weight: 700;">{base_n:,}</div>
                        </div>
                        <div style="background: white; border-radius: 12px; padding: 18px;
                             border: 1px solid {WB_BORDER}; text-align: center;">
                            <div style="color: {WB_MUTED}; font-size: 0.72rem; text-transform: uppercase;">Baseline Cum PD</div>
                            <div style="color: {WB_PRIMARY}; font-size: 1.5rem; font-weight: 700;">{base_cp_pct:.2f}%</div>
                        </div>
                        <div style="background: white; border-radius: 12px; padding: 18px;
                             border: 1px solid {WB_BORDER}; text-align: center;">
                            <div style="color: {WB_MUTED}; font-size: 0.72rem; text-transform: uppercase;">Baseline Ann. PD</div>
                            <div style="color: {WB_PRIMARY}; font-size: 1.5rem; font-weight: 700;">{base_ap_pct:.2f}%</div>
                        </div>
                        <div style="background: white; border-radius: 12px; padding: 18px;
                             border: 1px solid {WB_BORDER}; text-align: center;">
                            <div style="color: {WB_MUTED}; font-size: 0.72rem; text-transform: uppercase;">Target</div>
                            <div style="color: {WB_TEXT}; font-size: 1.5rem; font-weight: 700;">{target_pd_pct:.2f}%</div>
                        </div>
                        <div style="background: white; border-radius: 12px; padding: 18px;
                             border: 1px solid {WB_BORDER}; text-align: center;">
                            <div style="color: {WB_MUTED}; font-size: 0.72rem; text-transform: uppercase;">Gap to Target</div>
                            <div style="color: {'#10B981' if on_target else '#EF4444'}; font-size: 1.5rem; font-weight: 700;">{gap_pct:+.2f} pp</div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                # Show active constraints, if any
                if locked_cols or required_map:
                    parts = []
                    if locked_cols:
                        parts.append(f"🔒 Locked: {', '.join(locked_cols)}")
                    for c, rule in (required_map or {}).items():
                        if rule.get('type') == 'numeric':
                            parts.append(
                                f"📐 {c} must keep [{rule.get('lo'):.4g}, {rule.get('hi'):.4g}]"
                            )
                        else:
                            vals = rule.get('values') or []
                            preview = ', '.join(str(v) for v in vals[:5])
                            if len(vals) > 5:
                                preview += f', … (+{len(vals) - 5} more)'
                            parts.append(f"🏷️ {c} must include {{{preview}}}")
                    st.caption('Active constraints — ' + ' • '.join(parts))

                if on_target:
                    st.success(
                        f'✅ The current filtered portfolio is already at or below target '
                        f'({base_ap_pct:.2f}% ≤ {target_pd_pct:.2f}%). No tightening required.'
                    )
                elif np.isnan(base_ap_pct):
                    st.warning('Annualized PD could not be computed on the current portfolio '
                               '(insufficient observation time).')
                else:
                    st.info(
                        f'Current annualized PD is **{base_ap_pct:.2f}%** — '
                        f'need to reduce by **{gap_pct:.2f} pp** to reach '
                        f'**{target_pd_pct:.2f}%**.'
                    )

                    def _expand_categorical_lists(df: pd.DataFrame) -> pd.DataFrame:
                        """Add full Kept/Dropped value strings (categorical rows only) and
                        a Kept/Total count summary so the user can see every value."""
                        if df is None or df.empty:
                            return df
                        out = df.copy()

                        def _join(v):
                            if isinstance(v, list):
                                return ', '.join(str(x) for x in v)
                            return ''

                        out['Kept values (full)']    = out.get('Kept values', pd.Series([None] * len(out))).map(_join)
                        out['Dropped values (full)'] = out.get('Dropped values', pd.Series([None] * len(out))).map(_join)

                        kept_n  = out.get('Kept categories', pd.Series([None] * len(out)))
                        total_n = out.get('Total categories', pd.Series([None] * len(out)))

                        def _kt(k, t):
                            try:
                                return f'{int(k)} / {int(t)}'
                            except Exception:
                                return ''

                        out['Kept / Total'] = [_kt(k, t) for k, t in zip(kept_n, total_n)]
                        return out

                    def _render_table(df: pd.DataFrame, title: str, caption: str,
                                      gradient_col: str, gradient_cmap: str,
                                      download_key: str | None = None,
                                      download_name: str | None = None):
                        st.markdown(title)
                        st.caption(caption)
                        disp = _expand_categorical_lists(df)
                        show_cols = ['Column', 'Type', 'Filter',
                                     'Cum PD (%)', 'Annualized PD (%)',
                                     'Retained loans', 'Retained origination amt',
                                     'Retention (count) %', 'Retention (amt) %',
                                     'Kept / Total',
                                     'Kept values (full)', 'Dropped values (full)']
                        show_cols = [c for c in show_cols if c in disp.columns]
                        styled = (
                            disp[show_cols].style
                            .format({
                                'Cum PD (%)': '{:.2f}',
                                'Annualized PD (%)': '{:.2f}',
                                'Retained loans': '{:,.0f}' if pretty_ints else '{:.0f}',
                                'Retained origination amt': '{:,.2f}' if pretty_ints else '{:.2f}',
                                'Retention (count) %': '{:.1f}',
                                'Retention (amt) %':   '{:.1f}',
                            })
                            .background_gradient(subset=[gradient_col], cmap=gradient_cmap)
                            .hide(axis='index')
                        )
                        st.dataframe(styled, use_container_width=True, height=380)

                        if download_key and download_name:
                            # utf-8-sig BOM so Excel renders unicode (∈, ≤, ≥) correctly
                            csv_bytes = ('\ufeff' + disp[show_cols].to_csv(index=False)).encode('utf-8')
                            st.download_button(
                                '📊 Download recommendations (CSV)',
                                csv_bytes,
                                download_name,
                                'text/csv',
                                key=download_key,
                            )

                        # Per-row breakdown for categorical rows — full chip lists
                        cat_rows = disp[disp['Type'] == 'Categorical']
                        if not cat_rows.empty:
                            with st.expander(f'📋 Full kept / dropped category lists ({len(cat_rows)})',
                                             expanded=False):
                                for _, r in cat_rows.iterrows():
                                    kept = r.get('Kept values') if isinstance(r.get('Kept values'), list) else []
                                    dropped = r.get('Dropped values') if isinstance(r.get('Dropped values'), list) else []
                                    st.markdown(f"**{r['Column']}** — keep {len(kept)} of {len(kept) + len(dropped)} values "
                                                f"(Ann PD {r['Annualized PD (%)']:.2f}% · "
                                                f"retains {r['Retention (count) %']:.1f}% by count / "
                                                f"{r['Retention (amt) %']:.1f}% by amt)")
                                    cc1, cc2 = st.columns(2)
                                    with cc1:
                                        st.markdown('**✅ Keep**')
                                        if kept:
                                            st.markdown('\n'.join(f'- {v}' for v in kept))
                                        else:
                                            st.caption('— none —')
                                    with cc2:
                                        st.markdown('**❌ Exclude**')
                                        if dropped:
                                            st.markdown('\n'.join(f'- {v}' for v in dropped))
                                        else:
                                            st.caption('— none —')
                                    st.markdown('---')

                    if feasible_df is not None and not feasible_df.empty:
                        _render_table(
                            feasible_df,
                            title='#### 🟢 Single-filter recommendations that meet target',
                            caption='Ranked by retention (highest first). Each row is a standalone '
                                    'eligibility rule on top of the current sidebar filters.',
                            gradient_col='Retention (amt) %' if pd_by_amount_opt else 'Retention (count) %',
                            gradient_cmap='Greens',
                            download_key='opt_feasible_csv',
                            download_name=f'pd_optimizer_target_{target_pd_pct:.2f}pct.csv',
                        )
                    else:
                        st.warning(
                            '⚠️ No single attribute filter meets target while keeping at '
                            f'least {min_retention_pct}% retention. See the best-progress '
                            'table below and try the two-filter combo.'
                        )

                    if (feasible_df is None or feasible_df.empty) and partial_df is not None and not partial_df.empty:
                        _render_table(
                            partial_df,
                            title='#### 🟡 Best single-filter progress (does not reach target)',
                            caption='Per-column filter with the lowest achievable annualized PD '
                                    'subject to the minimum-retention constraint.',
                            gradient_col='Annualized PD (%)',
                            gradient_cmap='Blues_r',
                            download_key='opt_partial_csv',
                            download_name=f'pd_optimizer_progress_{target_pd_pct:.2f}pct.csv',
                        )

                    if combo is not None:
                        st.markdown('#### 🔗 Two-filter combo that meets target')
                        st.caption('Applies the two recommended filters together '
                                   '(AND logic) — often retains more than the most-aggressive single filter.')
                        combo_df = pd.DataFrame([combo])
                        show_cols = ['Filter 1', 'Filter 2', 'Cum PD (%)',
                                     'Annualized PD (%)', 'Retained loans',
                                     'Retained origination amt',
                                     'Retention (count) %', 'Retention (amt) %']
                        show_cols = [c for c in show_cols if c in combo_df.columns]
                        styled = (
                            combo_df[show_cols].style
                            .format({
                                'Cum PD (%)': '{:.2f}',
                                'Annualized PD (%)': '{:.2f}',
                                'Retained loans': '{:,.0f}' if pretty_ints else '{:.0f}',
                                'Retained origination amt': '{:,.2f}' if pretty_ints else '{:.2f}',
                                'Retention (count) %': '{:.1f}',
                                'Retention (amt) %':   '{:.1f}',
                            })
                            .hide(axis='index')
                        )
                        st.dataframe(styled, use_container_width=True, height=120)

                    st.markdown(
                        f"""
                        <div style="
                            background: {WB_LIGHT};
                            border-radius: 12px;
                            padding: 16px 20px;
                            border: 1px solid {WB_BORDER};
                            margin-top: 16px;
                            font-size: 0.85rem;
                            color: {WB_MUTED};
                        ">
                            <b style="color:{WB_TEXT};">How to read this</b><br/>
                            Numeric filters are reported as one-sided thresholds (≤ or ≥) that you can
                            map directly onto eligibility criteria in the sidebar. Categorical filters
                            keep the lowest-PD categories first; the "Kept categories" count tells you
                            how many of the original distinct values survive. Retention is measured
                            against the current sidebar-filtered portfolio.
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
            else:
                st.info(
                    '👆 Set a target annualized PD, choose the calculation basis, '
                    'and click **Generate recommendations** to scan every attribute '
                    'for the filter that meets target with the highest retention.'
                )

    else:
        # Empty state - no data loaded
        st.markdown(
            f"""
            <div style="
                background: white;
                border-radius: 20px;
                padding: 60px 40px;
                text-align: center;
                border: 2px dashed {WB_BORDER};
                margin: 40px 0;
            ">
                <div style="
                    background: {WB_LIGHT};
                    width: 80px;
                    height: 80px;
                    border-radius: 20px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    margin: 0 auto 24px auto;
                ">
                    <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="{WB_PRIMARY}" stroke-width="2">
                        <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>
                        <polyline points="17 8 12 3 7 8"/>
                        <line x1="12" y1="3" x2="12" y2="15"/>
                    </svg>
                </div>
                <h2 style="margin: 0 0 12px 0; color: {WB_TEXT}; font-weight: 600;">No Data Loaded</h2>
                <p style="margin: 0; color: {WB_MUTED}; font-size: 1rem; max-width: 400px; margin: 0 auto;">
                    Upload an Excel file (.xlsx) using the panel on the left to begin your vintage curve analysis.
                </p>
                <div style="margin-top: 24px;">
                    <span style="
                        background: {WB_LIGHT};
                        color: {WB_PRIMARY};
                        padding: 8px 16px;
                        border-radius: 20px;
                        font-size: 0.85rem;
                        font-weight: 500;
                    ">
                        Supported format: .xlsx
                    </span>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )
        















